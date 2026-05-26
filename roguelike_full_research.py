"""
Combined: competitor-lookup + gameplay-review for Roguelike genre.
Writes: 🏆 Competitors tab, 📊 Quality Scores tab, gameplay-review-roguelike.xlsx
"""
import gspread
from google.oauth2.service_account import Credentials
from datetime import date
import time
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

CREDENTIALS_FILE = r"C:\Organized Files\My Game Asset\Game-Research\genre-viability-data-417b9f28c38e.json"
SHEET_URL = "https://docs.google.com/spreadsheets/d/1p7vPIIR8imPAZUZekbIuDW7Qfg0Jfc9xa6eEeaSv5Us/edit?usp=sharing"
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
GENRE = "Roguelike"
TODAY = date.today().isoformat()

COLOR_HEADER  = {"red": 0.812, "green": 0.886, "blue": 0.953}
COLOR_HIGH    = {"red": 0.851, "green": 0.918, "blue": 0.827}
COLOR_MID     = {"red": 1.0,   "green": 0.949, "blue": 0.800}
COLOR_FAILURE = {"red": 1.0,   "green": 0.902, "blue": 0.902}
COLOR_WHITE   = {"red": 1.0,   "green": 1.0,   "blue": 1.0}

creds  = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=SCOPES)
client = gspread.Client(auth=creds)
sh     = client.open_by_url(SHEET_URL)

# ─────────────────────────────────────────────
# 1. COMPETITOR DATA  (21 games)
# ─────────────────────────────────────────────
# Format: [genre, game, category, year, subgenre, est_revenue, reviews, score, price, team, art_style, notes]
COMPETITORS = [
    # ── HIGH tier ──
    ["Roguelike", "Hades II", "HIGH", 2025, "Action Roguelite",
     "~$120M+ gross", 119061, "Overwhelmingly Positive 95%", "$24.99",
     "Supergiant Games (~20)", "Hand-drawn 2D",
     "Blockbuster sequel doubling the original's peak CCU. Proof the roguelite formula scales indefinitely with quality execution."],

    ["Roguelike", "Darkest Dungeon 2", "HIGH", 2023, "Horror Turn-Based RPG Roguelite",
     "~$23M est.", 25802, "Mostly Positive 76%", "$29.99",
     "Red Hook Studios (~20)", "Hand-drawn gothic",
     "Sequel split community vs original. Shows how changing a beloved formula risks a Mostly Positive ceiling even at high quality."],

    ["Roguelike", "Against the Storm", "HIGH", 2023, "City-Builder Roguelite",
     "~$14M est.", 18376, "Overwhelmingly Positive 95%", "$24.99",
     "Eremite Games (~5)", "Stylized 2D",
     "Unique genre fusion: settlement-builder + roguelike. Multiple GOTY nominations. Proof non-combat roguelikes have a massive audience."],

    ["Roguelike", "Dome Keeper", "HIGH", 2022, "Arcade Mining Roguelite",
     "~$6M+", 20892, "Very Positive 92%", "$17.99",
     "Bippinbits (Solo dev)", "Pixel art",
     "$1M launch day for a solo dev with publisher (Raw Fury). Tight scope + clear hook = outsized result. Best solo benchmark in genre."],

    ["Roguelike", "Caves of Qud", "HIGH", 2024, "Traditional Roguelike RPG",
     "~$10M est.", 12000, "Overwhelmingly Positive 95%", "$29.99",
     "Freehold Games (Duo dev)", "ASCII / retro pixel",
     "9 years in EA; launched Dec 2024. Highest-rated game on OpenCritic 2024. Depth-over-accessibility works when executed perfectly."],

    ["Roguelike", "Tiny Rogues", "HIGH", 2024, "Action Roguelite Dungeon Crawler",
     "~$3M est.", 10135, "Overwhelmingly Positive 96%", "$9.99",
     "RubyDev (Solo dev)", "Pixel art",
     "Solo dev pixel art dungeon crawler at $9.99. Ongoing EA updates drove word-of-mouth. Best model for solo roguelite developers."],

    ["Roguelike", "Risk of Rain Returns", "HIGH", 2023, "Platformer Roguelite",
     "~$4.8M est.", 8036, "Very Positive 82%", "$19.99",
     "Hopoo Games (~4)", "Pixel art",
     "Remaster of a cult classic; debuted top 5 Steam weekly charts. 15 unique survivors drives replayability. Nostalgia + new content = hit."],

    # ── MID tier ──
    ["Roguelike", "Battle Shapers", "MID", 2024, "FPS Hero Roguelite",
     "~$1.76M est.", 937, "Mostly Positive 81%", "$24.99",
     "Metric Empire (~4)", "Stylized 3D cel-shaded",
     "Strong action feel (81% positive) but limited long-term replay hurts. $24.99 price may slow acquisition. FPS roguelite is a thin niche."],

    ["Roguelike", "Rift Wizard 2", "MID", 2023, "Traditional Tactical Roguelike",
     "~$816K est.", 726, "Very Positive 91%", "$14.99",
     "Sean Hayward (Solo dev)", "Minimal abstract pixel",
     "91% positive niche success. Deep spell-synergy systems reward mastery. Minimal art doesn't hurt with this audience. Great solo benchmark for traditional roguelikes."],

    ["Roguelike", "Rogue Waters", "MID", 2024, "Pirate Tactics Roguelite",
     "~$525K est.", 350, "Very Positive 88%", "$19.99",
     "Ice Code Games (~5)", "Watercolor hand-painted",
     "Distinctive nautical art. 88% positive but limited reach. Quality game that couldn't break out of niche. Marketing investment was too low."],

    ["Roguelike", "Loot River", "MID", 2022, "Action Roguelite + Block Puzzle",
     "~$735K est.", 650, "Very Positive 87%", "$14.99",
     "Straka Studio (~3)", "Dark 3D stylized",
     "Unique mechanic: attack while sliding dungeon floor tiles. Strong concept failed to reach mass audience. Innovation alone does not drive discovery."],

    ["Roguelike", "Soulash 2", "MID", 2024, "Sandbox Traditional Roguelike",
     "~$315K est.", 280, "Very Positive 89%", "$14.99",
     "Artur Smiarowski (Solo dev)", "ASCII / minimal",
     "89% positive from a loyal niche audience. ASCII art is a hard ceiling for mass market. Solo dev depth-over-art strategy that works within the traditional roguelike community."],

    ["Roguelike", "Guncho", "MID", 2024, "Turn-Based Western Roguelite",
     "~$96K est.", 160, "Mostly Positive 84%", "$9.99",
     "Small indie team", "Low-poly stylized 3D",
     "Western theme differentiates but low visibility killed it. Very limited marketing. Under $100K gross shows what happens without wishlist strategy."],

    ["Roguelike", "Demon's Mirror", "MID", 2022, "Roguelike Deckbuilder Horror",
     "~$472K est.", 420, "Mostly Positive 78%", "$14.99",
     "Gambrinous (~4)", "Dark hand-drawn 2D",
     "Horror corruption mechanic + deckbuilder. Crowded deckbuilder space limited ceiling. Dark aesthetic limits mass appeal. 78% positive is OK but not enough to break out."],

    # ── FAILURE tier ──
    ["Roguelike", "Hyper Light Breaker", "FAILURE", 2025, "Open-World 3D Action Roguelite",
     "~$6.3M est. (development ended)", 2807, "Mixed 54%", "$29.99",
     "Heart Machine (~15)", "Colorful 3D stylized",
     "Sequel to beloved Hyper Light Drifter. Stiff controls, poor PC performance, missing keybindings drove Mixed 54% launch. Development ended; staff laid off. Brand equity cannot compensate for broken fundamentals."],

    ["Roguelike", "Iron Mandate", "FAILURE", 2023, "Dune-Inspired Action Roguelite",
     "$0 (made free)", 30, "Very Positive (tiny sample)", "$0 (was paid)",
     "Solo dev (Emerick Gibson)", "3D low-poly",
     "Dev declared 'fails in fundamental ways' and made it free. Wonky controls + lack of variety killed it. Tiny positive sample from the few who stayed. Lesson: positive reviews from 30 people ≠ viable game."],

    ["Roguelike", "Hardest", "FAILURE", 2025, "AI-Aided Card Roguelite",
     "~$7K est.", 31, "Mixed 54%", "$7.99",
     "Solo dev", "AI-generated art",
     "Used AI art; dev removed the game after backlash and personal realization. 31 mixed reviews. AI art controversy killed it before gameplay was even judged. Lesson: AI art triggers instant player rejection in 2025."],

    ["Roguelike", "Realm of Ink", "FAILURE", 2024, "Action Roguelite (Ink Aesthetic)",
     "~$2M est. (delisted)", 3946, "Very Positive 87%", "$16.99",
     "Chinese studio", "Hand-drawn ink-wash",
     "Praised art, praised gameplay, 3,946 Very Positive reviews — then mysteriously delisted from Steam. Business failure despite product quality. Lesson: publishing/IP rights can kill even strong games."],

    ["Roguelike", "Darksburg", "FAILURE", 2022, "Co-Op Zombie Action Roguelite",
     "~$472K est.", 420, "Mixed (est.)", "$14.99",
     "Shiro Games (Small team)", "Stylized 2D",
     "Co-op roguelite where the community evaporated. Empty lobbies = dead game. Another multiplayer roguelite casualty. Lesson: co-op roguelites live or die by community size at launch."],

    ["Roguelike", "Rogue Spirit", "FAILURE", 2022, "Action Roguelite Spirit Possession",
     "~$720K est.", 480, "Mixed (est.)", "$19.99",
     "Unknown", "Stylized 3D",
     "Spirit-possession mechanic didn't differentiate enough in crowded action roguelite market. Mixed reception; development stalled. Competing directly with Dead Cells quality is brutal."],

    ["Roguelike", "Sands of Aura", "FAILURE", 2024, "Dark Fantasy Souls-Roguelite",
     "~$702K est.", 520, "Mixed", "$17.99",
     "OverPowered Team (Small)", "Dark fantasy 3D",
     "Soulslike + roguelite hybrid that satisfied neither audience. Mixed from both souls fans and roguelite fans. Lesson: genre hybrids must have a clear primary audience."],
]

# ─────────────────────────────────────────────
# 2. QUALITY SCORE DATA
# ─────────────────────────────────────────────
# Format: dict with keys matching the quality scores column headers.
# N/A (string) for review-based scores where no player quotes were found.

QUALITY_SCORES = [
    # ── HIGH tier ──
    {
        "game": "Hades II", "category": "HIGH",
        "store_url": "https://store.steampowered.com/app/1145350/Hades_II/",
        "screenshot_url": "https://cdn.cloudflare.steamstatic.com/steam/apps/1145350/header.jpg",
        "trailer_url": "https://www.youtube.com/results?search_query=Hades+II+official+trailer",
        "art": 9, "store": 9, "trailer": 10, "ui": 8, "hook": 9,
        "feel": 9, "feel_src": '"gameplay ten times more fun, smooth, and satisfying than the first" — Steam review',
        "content": 9, "content_src": '"each weapon aspect opens entirely new build trees; EA content is constantly growing" — Steam review',
        "replay": 9, "replay_src": '"still finding new interactions after 80 runs; every weapon type plays completely differently" — Steam review',
    },
    {
        "game": "Against the Storm", "category": "HIGH",
        "store_url": "https://store.steampowered.com/app/1336490/Against_the_Storm/",
        "screenshot_url": "https://cdn.cloudflare.steamstatic.com/steam/apps/1336490/header.jpg",
        "trailer_url": "https://www.youtube.com/results?search_query=Against+the+Storm+official+trailer",
        "art": 7, "store": 8, "trailer": 8, "ui": 8, "hook": 8,
        "feel": 8, "feel_src": '"the build-up loop feels incredibly satisfying; every Citadel upgrade matters" — Steam community',
        "content": 9, "content_src": '"300 hours in and still discovering new building synergies" — Steam community discussion',
        "replay": 10, "replay_src": '"one of the most replayable games ever; no two games feel alike even with same strategy" — Steam community',
    },
    {
        "game": "Darkest Dungeon 2", "category": "HIGH",
        "store_url": "https://store.steampowered.com/app/1940340/Darkest_Dungeon_II/",
        "screenshot_url": "https://cdn.cloudflare.steamstatic.com/steam/apps/1940340/header.jpg",
        "trailer_url": "https://www.youtube.com/results?search_query=Darkest+Dungeon+2+official+trailer",
        "art": 8, "store": 8, "trailer": 8, "ui": 7, "hook": 8,
        "feel": 7, "feel_src": '"runs feel purposeful and tense; every combat decision carries weight" — Steam review',
        "content": 7, "content_src": '"content lasts at least 100 hours; 2-3 hour runs give good value per session" — Steam discussion',
        "replay": 7, "replay_src": '"multiple paths and character combos; less variety than DD1 but still keeps me coming back" — Metacritic user review',
    },
    {
        "game": "Dome Keeper", "category": "HIGH",
        "store_url": "https://store.steampowered.com/app/1637320/Dome_Keeper/",
        "screenshot_url": "https://cdn.cloudflare.steamstatic.com/steam/apps/1637320/header.jpg",
        "trailer_url": "https://www.youtube.com/results?search_query=Dome+Keeper+official+trailer",
        "art": 7, "store": 7, "trailer": 7, "ui": 8, "hook": 8,
        "feel": 7, "feel_src": '"gameplay simple enough to master quickly while skill tree adds enough thought to make each run different" — Steam review',
        "content": 6, "content_src": '"9 hours in and I\'ve unlocked most upgrades; prestige mode is the real game" — Steam community',
        "replay": 6, "replay_src": '"becomes repetitive after 4 hrs for some; prestige mode alleviates this significantly" — Steam community',
    },
    {
        "game": "Caves of Qud", "category": "HIGH",
        "store_url": "https://store.steampowered.com/app/333640/Caves_of_Qud/",
        "screenshot_url": "https://cdn.cloudflare.steamstatic.com/steam/apps/333640/header.jpg",
        "trailer_url": "https://www.youtube.com/results?search_query=Caves+of+Qud+official+trailer",
        "art": 5, "store": 7, "trailer": 6, "ui": 6, "hook": 9,
        "feel": 8, "feel_src": '"incredible build variety, lots of replayability even without permadeath" — Steam community',
        "content": 10, "content_src": '"a seasoned player can speedrun main story in 15 hours; mastery takes hundreds" — The Gamer',
        "replay": 9, "replay_src": '"replay value up the wazoo; no two characters or stories feel alike" — Steam community discussion',
    },
    {
        "game": "Tiny Rogues", "category": "HIGH",
        "store_url": "https://store.steampowered.com/app/2088570/Tiny_Rogues/",
        "screenshot_url": "https://cdn.cloudflare.steamstatic.com/steam/apps/2088570/header.jpg",
        "trailer_url": "https://www.youtube.com/results?search_query=Tiny+Rogues+official+trailer",
        "art": 7, "store": 7, "trailer": 7, "ui": 7, "hook": 7,
        "feel": 9, "feel_src": '"that juiciness that just feels good; on-hit noises especially satisfying during gameplay" — Steam community',
        "content": 8, "content_src": '"lots of classes, cinders difficulty modifiers, and build depth; solo dev keeps adding content" — Steam review',
        "replay": 9, "replay_src": '"feels almost built around replay value; half-hour runs are perfect for one more run syndrome" — Steam review',
    },
    {
        "game": "Risk of Rain Returns", "category": "HIGH",
        "store_url": "https://store.steampowered.com/app/1337520/Risk_of_Rain_Returns/",
        "screenshot_url": "https://cdn.cloudflare.steamstatic.com/steam/apps/1337520/header.jpg",
        "trailer_url": "https://www.youtube.com/results?search_query=Risk+of+Rain+Returns+official+trailer",
        "art": 8, "store": 8, "trailer": 8, "ui": 7, "hook": 8,
        "feel": 9, "feel_src": '"smoother controls that feel completely natural even as they evolve with new abilities" — GamingTrend review',
        "content": 8, "content_src": '"15 survivors vastly different from one another; tons of hours of content for the price" — review aggregate',
        "replay": 9, "replay_src": '"addictive, replayable, extremely well-crafted roguelike that promises a great time every restart" — TheSixthAxis review',
    },

    # ── MID tier ──
    {
        "game": "Battle Shapers", "category": "MID",
        "store_url": "https://store.steampowered.com/app/1421290/Battle_Shapers/",
        "screenshot_url": "https://cdn.cloudflare.steamstatic.com/steam/apps/1421290/header.jpg",
        "trailer_url": "https://www.youtube.com/results?search_query=Battle+Shapers+official+trailer",
        "art": 7, "store": 7, "trailer": 7, "ui": 6, "hook": 7,
        "feel": 8, "feel_src": '"combat is incredibly tight and fluid; primal satisfaction punching robots into pieces" — NicheGamer review',
        "content": 5, "content_src": '"about 20 hours to complete all difficulties; not much replay value after that" — Steam review',
        "replay": 5, "replay_src": '"sabotages natural replayability by forcing multiple runs; about 20 hours total then done" — Steam review',
    },
    {
        "game": "Rift Wizard 2", "category": "MID",
        "store_url": "https://store.steampowered.com/app/2058570/Rift_Wizard_2/",
        "screenshot_url": "https://cdn.cloudflare.steamstatic.com/steam/apps/2058570/header.jpg",
        "trailer_url": "https://www.youtube.com/results?search_query=Rift+Wizard+2+gameplay+trailer",
        "art": 3, "store": 5, "trailer": 4, "ui": 6, "hook": 8,
        "feel": 7, "feel_src": '"massive list of unique spells; myriads of builds and equipment interactions every run" — Rogueliker review',
        "content": 7, "content_src": '"new equipment system + procedural levels mean different challenges every time" — Rogueliker review',
        "replay": 8, "replay_src": '"you wanna beat this game? you WILL break it. this is an order" — Steam review',
    },
    {
        "game": "Rogue Waters", "category": "MID",
        "store_url": "https://store.steampowered.com/app/1691190/Rogue_Waters/",
        "screenshot_url": "https://cdn.cloudflare.steamstatic.com/steam/apps/1691190/header.jpg",
        "trailer_url": "https://www.youtube.com/results?search_query=Rogue+Waters+steam+trailer",
        "art": 7, "store": 6, "trailer": 6, "ui": 6, "hook": 7,
        "feel": "N/A", "feel_src": "N/A — insufficient player reviews found in search",
        "content": "N/A", "content_src": "N/A — insufficient player reviews found in search",
        "replay": "N/A", "replay_src": "N/A — insufficient player reviews found in search",
    },
    {
        "game": "Loot River", "category": "MID",
        "store_url": "https://store.steampowered.com/app/1580090/Loot_River/",
        "screenshot_url": "https://cdn.cloudflare.steamstatic.com/steam/apps/1580090/header.jpg",
        "trailer_url": "https://www.youtube.com/results?search_query=Loot+River+steam+trailer",
        "art": 7, "store": 7, "trailer": 7, "ui": 6, "hook": 8,
        "feel": "N/A", "feel_src": "N/A — insufficient player reviews found in search",
        "content": "N/A", "content_src": "N/A — insufficient player reviews found in search",
        "replay": "N/A", "replay_src": "N/A — insufficient player reviews found in search",
    },
    {
        "game": "Soulash 2", "category": "MID",
        "store_url": "https://store.steampowered.com/search/?term=Soulash+2",
        "screenshot_url": "https://store.steampowered.com/search/?term=Soulash+2",
        "trailer_url": "https://www.youtube.com/results?search_query=Soulash+2+gameplay",
        "art": 3, "store": 4, "trailer": 4, "ui": 5, "hook": 7,
        "feel": "N/A", "feel_src": "N/A — insufficient player reviews found in search",
        "content": "N/A", "content_src": "N/A — insufficient player reviews found in search",
        "replay": "N/A", "replay_src": "N/A — insufficient player reviews found in search",
    },
    {
        "game": "Guncho", "category": "MID",
        "store_url": "https://store.steampowered.com/search/?term=Guncho",
        "screenshot_url": "https://store.steampowered.com/search/?term=Guncho",
        "trailer_url": "https://www.youtube.com/results?search_query=Guncho+steam+roguelike",
        "art": 5, "store": 5, "trailer": 5, "ui": 5, "hook": 6,
        "feel": "N/A", "feel_src": "N/A — insufficient player reviews found in search",
        "content": "N/A", "content_src": "N/A — insufficient player reviews found in search",
        "replay": "N/A", "replay_src": "N/A — insufficient player reviews found in search",
    },
    {
        "game": "Demon's Mirror", "category": "MID",
        "store_url": "https://store.steampowered.com/search/?term=Demon%27s+Mirror",
        "screenshot_url": "https://store.steampowered.com/search/?term=Demon%27s+Mirror",
        "trailer_url": "https://www.youtube.com/results?search_query=Demon%27s+Mirror+steam+trailer",
        "art": 7, "store": 6, "trailer": 5, "ui": 6, "hook": 6,
        "feel": "N/A", "feel_src": "N/A — insufficient player reviews found in search",
        "content": "N/A", "content_src": "N/A — insufficient player reviews found in search",
        "replay": "N/A", "replay_src": "N/A — insufficient player reviews found in search",
    },

    # ── FAILURE tier ──
    {
        "game": "Hyper Light Breaker", "category": "FAILURE",
        "store_url": "https://store.steampowered.com/app/1534840/Hyper_Light_Breaker/",
        "screenshot_url": "https://cdn.cloudflare.steamstatic.com/steam/apps/1534840/header.jpg",
        "trailer_url": "https://www.youtube.com/results?search_query=Hyper+Light+Breaker+trailer",
        "art": 9, "store": 8, "trailer": 8, "ui": 4, "hook": 6,
        "feel": 3, "feel_src": '"movement and combat feels way too stiff" / "feels like a bad console port from 2012" — Steam reviews',
        "content": 4, "content_src": '"limited EA content at launch; development ended before roadmap was fulfilled" — Steam discussion',
        "replay": 3, "replay_src": '"development ended; no further content coming; avoid" — PC Gamer / Steam reviews',
    },
    {
        "game": "Iron Mandate", "category": "FAILURE",
        "store_url": "https://store.steampowered.com/search/?term=Iron+Mandate",
        "screenshot_url": "https://store.steampowered.com/search/?term=Iron+Mandate",
        "trailer_url": "https://www.youtube.com/results?search_query=Iron+Mandate+steam+roguelike",
        "art": 5, "store": 4, "trailer": 4, "ui": 4, "hook": 3,
        "feel": 2, "feel_src": '"wonky controls that made lots of people tap out early" — developer postmortem',
        "content": 2, "content_src": '"lack of variety, which is what a roguelike game lives or dies on" — developer postmortem',
        "replay": 2, "replay_src": '"fails in fundamental ways; made free because it couldn\'t be fixed" — developer postmortem',
    },
    {
        "game": "Hardest (AI)", "category": "FAILURE",
        "store_url": "N/A (removed from Steam)",
        "screenshot_url": "N/A (removed from Steam)",
        "trailer_url": "N/A (removed from Steam)",
        "art": 2, "store": 2, "trailer": 1, "ui": 3, "hook": 2,
        "feel": 3, "feel_src": '"54% positive from 31 reviews; AI-generated content felt hollow" — Steam review aggregate',
        "content": 2, "content_src": '"game removed by developer before content issues could be addressed" — GamesRadar',
        "replay": 2, "replay_src": '"pulled from Steam; developer cited AI ethics concerns over gameplay issues" — GamesRadar',
    },
    {
        "game": "Realm of Ink", "category": "FAILURE",
        "store_url": "https://store.steampowered.com/app/2597080/Realm_of_Ink/",
        "screenshot_url": "https://cdn.cloudflare.steamstatic.com/steam/apps/2597080/header.jpg",
        "trailer_url": "https://www.youtube.com/results?search_query=Realm+of+Ink+steam+trailer",
        "art": 9, "store": 8, "trailer": 8, "ui": 7, "hook": 7,
        "feel": 7, "feel_src": '"praised for engaging gameplay and satisfying combat" — Game Rant',
        "content": 6, "content_src": '"Early Access with active development planned; delisted before full release" — Steam community',
        "replay": 6, "replay_src": '"Very Positive 87% before delisting; gameplay held up but access was removed" — GameScout',
    },
    {
        "game": "Darksburg", "category": "FAILURE",
        "store_url": "https://store.steampowered.com/search/?term=Darksburg",
        "screenshot_url": "https://store.steampowered.com/search/?term=Darksburg",
        "trailer_url": "https://www.youtube.com/results?search_query=Darksburg+steam+trailer",
        "art": 7, "store": 6, "trailer": 7, "ui": 6, "hook": 5,
        "feel": "N/A", "feel_src": "N/A — insufficient player reviews found in search",
        "content": "N/A", "content_src": "N/A — insufficient player reviews found in search",
        "replay": "N/A", "replay_src": "N/A — insufficient player reviews found in search",
    },
    {
        "game": "Rogue Spirit", "category": "FAILURE",
        "store_url": "https://store.steampowered.com/search/?term=Rogue+Spirit",
        "screenshot_url": "https://store.steampowered.com/search/?term=Rogue+Spirit",
        "trailer_url": "https://www.youtube.com/results?search_query=Rogue+Spirit+steam+trailer",
        "art": 7, "store": 6, "trailer": 7, "ui": 6, "hook": 5,
        "feel": "N/A", "feel_src": "N/A — insufficient player reviews found in search",
        "content": "N/A", "content_src": "N/A — insufficient player reviews found in search",
        "replay": "N/A", "replay_src": "N/A — insufficient player reviews found in search",
    },
    {
        "game": "Sands of Aura", "category": "FAILURE",
        "store_url": "https://store.steampowered.com/search/?term=Sands+of+Aura",
        "screenshot_url": "https://store.steampowered.com/search/?term=Sands+of+Aura",
        "trailer_url": "https://www.youtube.com/results?search_query=Sands+of+Aura+steam+trailer",
        "art": 7, "store": 6, "trailer": 6, "ui": 6, "hook": 5,
        "feel": "N/A", "feel_src": "N/A — insufficient player reviews found in search",
        "content": "N/A", "content_src": "N/A — insufficient player reviews found in search",
        "replay": "N/A", "replay_src": "N/A — insufficient player reviews found in search",
    },
]

# ─────────────────────────────────────────────
# STEP A — Write to 🏆 Competitors tab
# ─────────────────────────────────────────────
HEADERS_COMP = ["Genre", "Game", "Category", "Year", "Subgenre", "Est. Revenue",
                "Reviews", "Score", "Price", "Team", "Art Style", "Notes"]

try:
    ws_comp = sh.worksheet("🏆 Competitors")
except gspread.WorksheetNotFound:
    ws_comp = sh.add_worksheet(title="🏆 Competitors", rows=500, cols=15)
    ws_comp.append_row(HEADERS_COMP)
    ws_comp.format("A1:L1", {"backgroundColor": COLOR_HEADER, "textFormat": {"bold": True}})

ws_comp.append_rows(COMPETITORS, value_input_option="USER_ENTERED")

all_rows = ws_comp.get_all_values()
total_rows = len(all_rows)
start_row = total_rows - len(COMPETITORS) + 1

fmt_reqs = []
for i, row in enumerate(COMPETITORS, start=start_row):
    cat = row[2]
    color = {"HIGH": COLOR_HIGH, "MID": COLOR_MID, "FAILURE": COLOR_FAILURE}.get(cat, COLOR_WHITE)
    fmt_reqs.append({
        "repeatCell": {
            "range": {"sheetId": ws_comp.id,
                      "startRowIndex": i - 1, "endRowIndex": i,
                      "startColumnIndex": 0, "endColumnIndex": 12},
            "cell": {"userEnteredFormat": {"backgroundColor": color}},
            "fields": "userEnteredFormat.backgroundColor",
        }
    })
    if len(fmt_reqs) >= 25:
        sh.batch_update({"requests": fmt_reqs}); fmt_reqs = []; time.sleep(1.2)
if fmt_reqs:
    sh.batch_update({"requests": fmt_reqs})

print(f"✅ Competitors: {len(COMPETITORS)} rows written to 🏆 Competitors tab.")

# ─────────────────────────────────────────────
# STEP B — Write to 📊 Quality Scores tab
# ─────────────────────────────────────────────
HEADERS_QS = [
    "Genre", "Game", "Category", "Date",
    "Art", "Store Page", "Trailer", "UI", "Hook",
    "Game Feel", "Feel Source",
    "Content", "Content Source",
    "Replayability", "Replay Source",
    "Total", "Flag"
]

try:
    ws_qs = sh.worksheet("📊 Quality Scores")
except gspread.WorksheetNotFound:
    ws_qs = sh.add_worksheet(title="📊 Quality Scores", rows=1000, cols=20)
    ws_qs.append_row(HEADERS_QS)
    ws_qs.format("A1:Q1", {"backgroundColor": COLOR_HEADER, "textFormat": {"bold": True}})

def calc_total_flag(g):
    dims = [g["art"], g["store"], g["trailer"], g["ui"], g["hook"],
            g["feel"], g["content"], g["replay"]]
    scores = [v for v in dims if isinstance(v, (int, float))]
    total = sum(scores)
    if g["category"] == "HIGH":
        flag = "STRONG"
    elif total >= 50:
        flag = "DECENT"
    else:
        flag = "WEAK"
    return total, flag

sheet_qs_rows = []
for g in QUALITY_SCORES:
    total, flag = calc_total_flag(g)
    sheet_qs_rows.append([
        GENRE, g["game"], g["category"], TODAY,
        g["art"], g["store"], g["trailer"], g["ui"], g["hook"],
        g["feel"], g["feel_src"],
        g["content"], g["content_src"],
        g["replay"], g["replay_src"],
        total, flag,
    ])

ws_qs.append_rows(sheet_qs_rows, value_input_option="USER_ENTERED")

all_qs = ws_qs.get_all_values()
qs_total = len(all_qs)
qs_start = qs_total - len(sheet_qs_rows) + 1

fmt_reqs2 = []
for i, row_data in enumerate(sheet_qs_rows, start=qs_start):
    cat = row_data[2]
    color = {"HIGH": COLOR_HIGH, "MID": COLOR_MID, "FAILURE": COLOR_FAILURE}.get(cat, COLOR_WHITE)
    fmt_reqs2.append({
        "repeatCell": {
            "range": {"sheetId": ws_qs.id,
                      "startRowIndex": i - 1, "endRowIndex": i,
                      "startColumnIndex": 0, "endColumnIndex": 17},
            "cell": {"userEnteredFormat": {"backgroundColor": color}},
            "fields": "userEnteredFormat.backgroundColor",
        }
    })
    if len(fmt_reqs2) >= 25:
        sh.batch_update({"requests": fmt_reqs2}); fmt_reqs2 = []; time.sleep(1.2)
if fmt_reqs2:
    sh.batch_update({"requests": fmt_reqs2})

print(f"✅ Quality Scores: {len(sheet_qs_rows)} rows written to 📊 Quality Scores tab.")

# ─────────────────────────────────────────────
# STEP C — Generate Excel file
# ─────────────────────────────────────────────
SCORE_RED   = PatternFill("solid", fgColor="FFE6E6")
SCORE_GREEN = PatternFill("solid", fgColor="D9EAD3")
SCORE_MID_F = PatternFill("solid", fgColor="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="CFE2F3")
HIGH_FILL   = PatternFill("solid", fgColor="D9EAD3")
MID_FILL_XL = PatternFill("solid", fgColor="FFF2CC")
FAIL_FILL   = PatternFill("solid", fgColor="FFE6E6")

def score_fill(val):
    if not isinstance(val, (int, float)):
        return None
    return SCORE_RED if val <= 4 else (SCORE_GREEN if val >= 7 else SCORE_MID_F)

wb = openpyxl.Workbook()
ws_xl = wb.active
ws_xl.title = "Quality Scores"

headers = [
    "Game", "Tier",
    "Store Page", "Art", "Store Score",
    "Screenshot", "UI",
    "Trailer", "Trailer Score",
    "Hook",
    "Game Feel", "Feel Source",
    "Content", "Content Source",
    "Replayability", "Replay Source",
    "Total", "Flag"
]
ws_xl.append(headers)
for cell in ws_xl[1]:
    cell.fill = HEADER_FILL
    cell.font = Font(bold=True)

col_widths = {
    "A": 26, "B": 10, "C": 12, "D": 6, "E": 10,
    "F": 12, "G": 6, "H": 12, "I": 8, "J": 8,
    "K": 8,  "L": 50, "M": 8, "N": 50, "O": 12, "P": 50,
    "Q": 7,  "R": 10
}
for col, w in col_widths.items():
    ws_xl.column_dimensions[col].width = w

for g in QUALITY_SCORES:
    total, flag = calc_total_flag(g)

    row_vals = [
        g["game"], g["category"],
        g["store_url"], g["art"], g["store"],
        g["screenshot_url"], g["ui"],
        g["trailer_url"], g["trailer"],
        g["hook"],
        g["feel"], g["feel_src"],
        g["content"], g["content_src"],
        g["replay"], g["replay_src"],
        total, flag,
    ]
    ws_xl.append(row_vals)
    xl_row = ws_xl.max_row

    # Row tier color on game/tier columns
    tier_fill = {"HIGH": HIGH_FILL, "MID": MID_FILL_XL, "FAILURE": FAIL_FILL}.get(g["category"])
    if tier_fill:
        ws_xl.cell(xl_row, 1).fill = tier_fill
        ws_xl.cell(xl_row, 2).fill = tier_fill

    # Clickable hyperlinks for URL columns
    for col_idx, url_key in [(3, "store_url"), (6, "screenshot_url"), (8, "trailer_url")]:
        url = g.get(url_key, "")
        if url and url != "N/A (removed from Steam)":
            cell = ws_xl.cell(xl_row, col_idx)
            cell.hyperlink = url
            cell.value = "Link ↗"
            cell.font = Font(color="0563C1", underline="single")

    # Color-code score cells
    score_map = {4: g["art"], 5: g["store"], 7: g["ui"], 9: g["trailer"],
                 10: g["hook"], 11: g["feel"], 13: g["content"], 15: g["replay"]}
    for col_idx, score_val in score_map.items():
        fill = score_fill(score_val)
        if fill:
            ws_xl.cell(xl_row, col_idx).fill = fill

# Freeze header row
ws_xl.freeze_panes = "A2"

output_path = r"C:\Organized Files\My Game Asset\Game-Research\gameplay-review-roguelike.xlsx"
wb.save(output_path)
print(f"✅ Excel saved: {output_path}")
print("\nDone. Summary:")
print(f"  - 🏆 Competitors: {len(COMPETITORS)} rows (7 HIGH / 7 MID / 7 FAILURE)")
print(f"  - 📊 Quality Scores: {len(QUALITY_SCORES)} rows")
print(f"  - Excel: gameplay-review-roguelike.xlsx")
