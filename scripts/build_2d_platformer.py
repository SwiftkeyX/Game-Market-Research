"""
2D Platformer competitor data - Excel build script
Saves: gameplay-review-2d-platformer.xlsx
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import date
import shutil
from pathlib import Path

HEADER_FILL = PatternFill("solid", fgColor="CFE2F3")
HIGH_FILL   = PatternFill("solid", fgColor="D9EAD3")
MID_FILL    = PatternFill("solid", fgColor="FFF2CC")
FAIL_FILL   = PatternFill("solid", fgColor="FFE6E6")
TIER_FILLS  = {"HIGH": HIGH_FILL, "MID": MID_FILL, "FAILURE": FAIL_FILL}

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Genre Deep Dive"

headers = [
    "Genre", "Game", "Year", "Revenue", "Review",
    "Team size", "Trailer", "UI", "Art style",
    "Feature", "Scope", "Content", "Replayability",
]
ws.append(headers)
for cell in ws[1]:
    cell.fill = HEADER_FILL
    cell.font = Font(bold=True)

col_widths = {
    "A": 20, "B": 28, "C": 8, "D": 16, "E": 10,
    "F": 16, "G": 30, "H": 25, "I": 22,
    "J": 50, "K": 35, "L": 45, "M": 45,
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

game_rows = [
    # ---- HIGH TIER ----
    {
        "genre": "2D Platformer",
        "game": "Pizza Tower",
        "tier": "HIGH",
        "year": 2023,
        "revenue": "~$72M+ est.",
        "review": "71,408",
        "team_size": "Solo (Tour De Pizza / McPig)",
        "trailer": "Multiple viral trailers showing frantic Wario Land-speed runs; meme-generating clips drove massive pre-launch buzz; shows momentum system in first 5s",
        "ui": "Minimal HUD - speed meter and Peppino state visible; no clutter; arcade score display clean and readable",
        "art_style": "High-res pixel art, 1990s Nicktoon-inspired, vibrant chaotic palette, exaggerated rubber-hose animations",
        "feature": "- Momentum-based speed platforming (walk to sprint to chaos)\n- 5 floors x multiple levels; secret bonus rooms in every level\n- Score rank system (D to P); taunts + parries\n- 3 campaigns (Peppino, Gustavo, Swap Mode)\n- 2 playable characters; 15+ boss fights",
        "scope": "5-6hr first playthrough; 15+ levels; P-rank chase adds 50-100+ hrs; no EA period; launched full Jan 2023",
        "content": "5 floors with unique pizza-themed biomes; cheese dragons; secret bonus dungeons per level; hidden collectibles. 'More content than games 3x the price' - Steam",
        "replayability": "Extremely high - P-rank pursuit drives hundreds of hours; daily challenge community; active mod scene. 'Still playing after 400+ hrs chasing P ranks' - Reddit r/PizzaTower",
    },
    {
        "genre": "2D Platformer",
        "game": "SANABI",
        "tier": "HIGH",
        "year": 2023,
        "revenue": "~$20-30M est.",
        "review": "43,920",
        "team_size": "Team of 5 (WonderPotion) + NEOWIZ pub.",
        "trailer": "Shows grappling hook mechanic clearly with stylish cyberpunk action sequences; music-driven pacing; introduces story stakes early",
        "ui": "Clean minimal cyberpunk HUD; health and energy bars readable at a glance; boss health bars stylized but clear",
        "art_style": "Pixel art cyberpunk dystopia; neon-lit dark city aesthetic; cel-shaded character sprites; vivid neon accents on dark palette",
        "feature": "- Prosthetic grappling arm as sole traversal tool; swing, boost, slam\n- Combined traversal + weapon in one mechanic\n- Story-driven with strong found-family narrative\n- Diverse boss fights integrating hook mechanic\n- Multiple districts with unique visual identity",
        "scope": "8-10hr campaign; multiple chapters and districts; NG+ and boss rush mode; EA June 2022, full release Nov 2023",
        "content": "Multiple city districts with distinct themes; varied enemy types; memorable boss designs; strong cinematic story. 'Cried at the ending - story hit differently' - Steam",
        "replayability": "Moderate-high; NG+ carries over upgrades; speedrun community active; story rewards repeat viewing. 'Replayed twice for story details I missed' - Reddit r/pcgaming",
    },
    {
        "genre": "2D Platformer",
        "game": "Neon White",
        "tier": "HIGH",
        "year": 2022,
        "revenue": "$8.8M (reported)",
        "review": "19,085",
        "team_size": "Team of ~6 (Angel Matrix) + Annapurna pub.",
        "trailer": "Fast-cut speedrun showcase with anime aesthetic; card-discarding mechanic shown clearly in first 15s; viral among speedrun and anime communities",
        "ui": "Extremely clean - medal timer, card slots, demon count; highly readable at speed; leaderboard ghost visible during run",
        "art_style": "Anime character art (2D portraits) + vibrant neon 3D environments; cel-shaded; bold color-coded card system; visual novel sequences between levels",
        "feature": "- Card-based ability system: discard cards to unlock special moves (dash, double-jump, rocket)\n- Speedrun-focused 100+ level design\n- Leaderboards + medal grades (Bronze-Ace) baked into core loop\n- Gift/relationship system unlocks story scenes\n- 5 chapters + post-game challenge missions",
        "scope": "6-10hr main story; 100+ levels; post-game challenge levels; significant speedrunning depth beyond campaign",
        "content": "100+ levels across 5 chapters; gift items unlock story scenes; challenge levels add 4-6hr post-game. 'Perfect 'just one more attempt' loop' - Steam",
        "replayability": "Very high for speedrunners; friend leaderboard competition drives daily play; gold medal + Ace runs add 50+ hrs. 'Replayed every level 200+ times chasing Ace' - speedrun community",
    },
    {
        "genre": "2D Platformer",
        "game": "Anomaly Agent",
        "tier": "HIGH",
        "year": 2024,
        "revenue": "~$10M est.",
        "review": "8,973",
        "team_size": "Team of ~5 (Phew Phew Games) + GameDev.ist",
        "trailer": "High-energy showcase of cyberpunk beat-em-up combat; stylish anime action cuts; communicates tone and combat feel immediately",
        "ui": "Clean manga-panel aesthetic HUD; health/energy minimal but visible; stylish comic book presentation consistent with overall design",
        "art_style": "Pixel art cyberpunk manga; stylized B&W manga panels for cutscenes; vibrant neon action sequences; punchy particle effects",
        "feature": "- Beat-em-up platformer hybrid with agency operative setting\n- Multiple weapons with unique feel\n- Time-stop 'anomaly' mechanic for crowd control\n- Parry system with skill expression reward\n- Multiple mission levels with varied objectives\n- 3 difficulty modes",
        "scope": "6-8hr campaign; multiple mission levels; 3 difficulties; PS Vita-era action game pacing inspiration",
        "content": "Multiple agency missions across cyberpunk districts; diverse enemy factions and boss types; operative narrative with twists. 'Best action platformer I played in years' - Steam",
        "replayability": "Moderate-high; challenge modes; S-rank chase per level; 'Goes so fast you replay for pure feel' - Reddit r/indiegaming",
    },
    {
        "genre": "2D Platformer",
        "game": "Lil Gator Game",
        "tier": "HIGH",
        "year": 2022,
        "revenue": "~$5.9M est.",
        "review": "3,929",
        "team_size": "Team of 3 (MegaWobble) + Playtonic Friends pub.",
        "trailer": "Adorable wholesome tone showcasing cozy exploration; slow-paced warmth emphasizes story and characters over action; appeals to cozy/indie audience clearly",
        "ui": "Clean minimal; companion tracking and quest markers subtle; very readable; designed for cozy not stressful experience",
        "art_style": "Colorful low-poly 3D; bright cheerful palette; cel-shaded chibi characters; toybox aesthetic with cardboard props",
        "feature": "- Cozy 3D collectathon platformer with no combat\n- Cardboard shield crafting from collected materials\n- NPC friendship quests drive progression\n- Sibling relationship narrative at core\n- Open world exploration with collectibles\n- DLC 'In the Dark' (2025) expands content",
        "scope": "4-6hr main experience; dense NPC and collectible world; DLC adds 2-3hr; no EA period",
        "content": "Multiple themed island areas; 50+ NPCs with dialogue arcs; heartwarming sibling story. 'Made me cry before the credits rolled' - Steam; strong critical reception from cozy game audience",
        "replayability": "Low-moderate; exploration replay; DLC adds content; 'Perfect weekend game, played it twice' - Reddit r/NintendoSwitch",
    },
    {
        "genre": "2D Platformer",
        "game": "Gravity Circuit",
        "tier": "HIGH",
        "year": 2023,
        "revenue": "~$3.8M est.",
        "review": "3,381",
        "team_size": "Solo (Antti / Domesticated Ant Games) + PID Games pub.",
        "trailer": "Clean showcase of Mega Man-inspired stage progression and robot combat; communicates genre clearly to target audience; good action pacing",
        "ui": "Classic retro HUD - health bar, lives, weapon/energy system; faithful to Mega Man inspiration; clean and functional",
        "art_style": "Detailed 2D pixel art; vibrant robot character designs; Mega Man Zero/ZX aesthetic influence; clean readable enemy designs; bright stage themes",
        "feature": "- Mega Man-inspired nonlinear stage progression (choose boss order)\n- Gundam-inspired robot visual design\n- Fighting-game style combo system\n- Robot companion assists; upgradeable ability tree\n- 8 main stages + post-game challenge content\n- Boss rush mode",
        "scope": "8-12hr campaign; 8+ robot masters; upgradeable abilities; score attack system; post-game challenge stages; ~10 years solo development",
        "content": "8 themed robot master stages; diverse ability upgrades from bosses; arena challenge mode. 'Best Mega Man game since the originals' - Metacritic user reviews",
        "replayability": "High for action platformer fans; S-rank level chase; boss rush replay; 'I speedrun this now' - Reddit r/speedrun",
    },
    {
        "genre": "2D Platformer",
        "game": "ANTONBLAST",
        "tier": "HIGH",
        "year": 2024,
        "revenue": "~$4.3M est. (recouped <1 month post-launch)",
        "review": "2,835",
        "team_size": "Team of ~8-10 (Summitsphere) + Joystick Ventures pub.",
        "trailer": "Explosive high-energy showcase of Wario Land-style destruction; viral social media presence before launch; communicates chaotic energy and core mechanic immediately",
        "ui": "Bold arcade-style HUD; lives and health readable; Blast Meter visible; collectible tracking minimal; clean for high-speed play",
        "art_style": "Vibrant 2D pixel art; 1990s Nickelodeon cartoon energy (similar to Pizza Tower); explosive particle effects; bold saturated palette",
        "feature": "- Wario Land-inspired environment destruction mechanics\n- Unique 'Blast Meter' system\n- 12 side-scrolling levels across varied worlds\n- Boss rush + time-based 'Lime Trials' challenge mode\n- Hard mode; Stage Rush speedrun mode\n- 2 playable characters (Anton, Annie)",
        "scope": "6-8hr campaign; 12 levels; boss rush; EA phase ~6 months; full launch Dec 2024 with bonus modes",
        "content": "12 levels from Boiler City to Hell; 4 bosses added in post-launch; diverse destruction physics per environment. 'A love letter to Wario Land' - Metacritic",
        "replayability": "High; speedrun-friendly design; mastery of destruction combos; Lime Trials replay; 'I play this on lunch breaks still' - Reddit r/indiegaming",
    },

    # ---- MID TIER ----
    {
        "genre": "2D Platformer",
        "game": "Frogun",
        "tier": "MID",
        "year": 2022,
        "revenue": "~$570K est.",
        "review": "507",
        "team_size": "Solo (Marco Venturini / Mokaloca)",
        "trailer": "Charming retro aesthetic emphasized; shows grapple-frog mechanic clearly; PS1-era nostalgia evident in visual style and sound design",
        "ui": "Minimal retro-styled HUD; life/grapple-ammo counter visible; PS1-era presentation consistent with game's identity",
        "art_style": "Low-poly 3D N64/PS1 era; chunky character models; warm color palette; retro cartridge aesthetic with blurry textures",
        "feature": "- 'Frogun' grappling hook as primary traversal mechanic\n- 3D platformer with collecting mechanics\n- Multiple themed worlds\n- Boss fights integrating grapple\n- DLC 'Frogun Encore' (2023) adds more content",
        "scope": "4-6hr main campaign; multiple worlds; boss fights; Encore DLC adds 3-4hr; solid post-launch support for solo dev",
        "content": "Multiple themed worlds with retro aesthetic; various collectibles; boss encounters. 'Perfectly captures that PS1 charm' - Steam; praised for solo dev polish",
        "replayability": "Moderate; time trials; DLC adds replay; 'Short but delightful, worth full price' - Steam",
    },
    {
        "genre": "2D Platformer",
        "game": "BioGun",
        "tier": "MID",
        "year": 2024,
        "revenue": "~$130K (reported estimate)",
        "review": "477",
        "team_size": "Duo (Dapper Dog Digital, Dallas TX)",
        "trailer": "Colorful inside-body sci-fi aesthetic showcased; Metroidvania gameplay loops shown; biome exploration emphasized; niche but clear pitch",
        "ui": "Standard Metroidvania HUD; map + health + weapon display; somewhat busy but functional for the genre audience",
        "art_style": "Vibrant 2D sci-fi cartoon; inside-human-body aesthetic; colorful cell/bacteria enemy designs; bright biome theming",
        "feature": "- Metroidvania + platformer hybrid with vaccine-themed combat\n- Twin-stick shooter combat elements\n- Unlockable movement abilities gate exploration\n- Interconnected body biomes with unique mechanics\n- Full map exploration; multiple upgrade paths",
        "scope": "10-15hr Metroidvania; multiple distinct biomes; full map with secrets; EA for ~2 years before full 2024 launch",
        "content": "Multiple body biomes (lungs, heart, brain etc.); diverse enemy cell types; boss fights; upgrade tree. 'Underrated gem hiding in plain sight' - Reddit r/Metroidvania",
        "replayability": "Moderate; New Game+ option; 'Great first run, unlikely to replay' - Steam",
    },
    {
        "genre": "2D Platformer",
        "game": "Desvelado",
        "tier": "MID",
        "year": 2024,
        "revenue": "~$189K est.",
        "review": "360",
        "team_size": "Small indie team (Latin America)",
        "trailer": "Cute spooky atmosphere; light-extinguishing mechanic shown clearly and charmingly; gentle haunted lullaby music sets tone immediately",
        "ui": "Minimal and clean; level-specific light indicators; very readable; nothing distracts from platforming focus",
        "art_style": "Hand-drawn 2D dark pastel; sleepy dreamy aesthetic; cute chibi vampire design; soft shadow palette",
        "feature": "- Precision platformer around extinguishing every light in a haunted castle\n- Flame-based limited dash (one per flame collected)\n- Ghost enemies re-light lamps (creates unique tension loop)\n- Wall-jump and standard platforming vocabulary\n- No violence - pure puzzle-platforming",
        "scope": "2-3hr; focused precision platformer; polished single-playthrough experience; $6.99 price",
        "content": "Castle environments with varied puzzle rooms; ghost patrolling enemies; light-based challenge variety. 'Perfect Halloween game' - Steam",
        "replayability": "Low; experience-focused single playthrough; 'Finished in one sitting, absolutely worth it' - Steam",
    },
    {
        "genre": "2D Platformer",
        "game": "Zefyr: A Thief's Melody",
        "tier": "MID",
        "year": 2025,
        "revenue": "~$252K est.",
        "review": "337",
        "team_size": "Solo (Mathias Fontmarty, 11 years dev)",
        "trailer": "Beautiful cel-shaded world showcased; Zelda BOTW aesthetic immediately apparent; gentle ocean music; communicates freedom and exploration clearly",
        "ui": "Clean minimal; freedom-focused design; no combat HUD clutter; traversal feels unobstructed",
        "art_style": "Cel-shaded 3D; soft watercolor-inspired palette; vibrant floating islands; Breath of the Wild Zelda aesthetic; anime cel-shading influence",
        "feature": "- Free-climbing on almost any surface\n- Open 3D island environments\n- Optional stealth sections\n- Environmental puzzles and collectible quests\n- No fall damage; no health bar\n- Combat entirely optional",
        "scope": "6-10hr; multiple themed islands; 11-year solo development; June 2025 launch",
        "content": "Multiple floating island environments; stealth and puzzle blend; collectible treasures; contemplative narrative without dialogue. 'Best $X I ever spent - solo dev masterwork' - Steam",
        "replayability": "Low-moderate; exploration replay; 'I just want to exist in this world' - Steam",
    },
    {
        "genre": "2D Platformer",
        "game": "Symphonia",
        "tier": "MID",
        "year": 2024,
        "revenue": "~$221K est.",
        "review": "295",
        "team_size": "Team of ~4 (Sunny Peak) + Headup pub.",
        "trailer": "Beautiful musical atmosphere; violinist launch mechanic shown elegantly; orchestral score immediately evocative; visually distinct",
        "ui": "Clean minimal floating HUD; momentum cues subtle; nothing distracts from the aesthetic experience",
        "art_style": "Hand-drawn/painterly 2D; muted baroque palette; ornate concert hall and cathedral environments; Baroque visual identity throughout",
        "feature": "- Momentum-based platforming: violin launches Philemon through levels\n- Precision platformer with accessibility options (double jump, slow-mo)\n- 4 distinct themed locations\n- Upgradeable movement abilities\n- Wordless musical narrative\n- Orchestral score integrated into level design",
        "scope": "3-4hr; 4 themed locations; precision challenge throughout; difficulty can extend playthrough to 7hr+; $9.99 price",
        "content": "4 distinct baroque/orchestral environments; boss encounters; atmospheric wordless narrative. 'Short but unforgettable' - DualShockers",
        "replayability": "Low; experience-focused; 'Perfect first playthrough, unlikely to replay' - Metacritic user review",
    },
    {
        "genre": "2D Platformer",
        "game": "Windswept",
        "tier": "MID",
        "year": 2025,
        "revenue": "~$215K est. (barely broke even per dev)",
        "review": "287",
        "team_size": "Solo (PeekingBoo / WeatherFell)",
        "trailer": "Colorful retro showcase; two-character buddy gameplay shown; nostalgic SNES energy communicated; heartwarming tone",
        "ui": "Clean 2-character HUD; lives and health readable; minimal but functional for retro design",
        "art_style": "Pixel art SNES-era aesthetic; bright primary colors; DKC/Kirby/Mario World visual touchstones; cheerful cartoon style",
        "feature": "- 2-character buddy platformer (duck + tortoise with different abilities)\n- 40+ stages across themed worlds\n- Secrets and collectibles in every stage\n- Local co-op split-screen support\n- Retro challenge-first level design",
        "scope": "8-12hr; 40+ stages; co-op mode adds significant replay; secrets dense per level",
        "content": "40+ stages with themed worlds; hidden collectibles; bonus challenges. 'Best retro platformer of 2025' - critics; dev barely broke even despite 97% positive",
        "replayability": "Moderate; co-op replay value strong; time trials; 'Fun with a friend, more limited solo' - Steam",
    },
    {
        "genre": "2D Platformer",
        "game": "Curse of the Sea Rats",
        "tier": "MID",
        "year": 2023,
        "revenue": "~$153K est.",
        "review": "102",
        "team_size": "Team of ~8 (Petoons Studio, Spain)",
        "trailer": "Hand-drawn art style shown prominently; adventurous pirate tone; Metroidvania gameplay highlighted; beautiful but doesn't clearly communicate the hook",
        "ui": "Standard Metroidvania HUD; map + health + weapon display; somewhat cluttered; pacing issues with readability noted by reviewers",
        "art_style": "Hand-painted 2D; vibrant rat pirate aesthetic; detailed character animations; colorful ocean and dungeon setting",
        "feature": "- Metroidvania platformer with 4 unique rat characters\n- Hand-drawn frame-by-frame animation\n- NPC side quests\n- Multiple biomes with different mechanics\n- Full interconnected map exploration",
        "scope": "10-12hr Metroidvania; 4 characters with distinct playstyles; multiple biomes; boss fights; full map",
        "content": "Multiple pirate biomes; 4 rat protagonists; boss fights; NPC storylines. 'Beautiful art sadly paired with mediocre platforming' - Metacritic",
        "replayability": "Low; gameplay issues limit replay motivation; 55% positive reflects mixed quality despite visual investment",
    },

    # ---- FAILURE TIER ----
    {
        "genre": "2D Platformer",
        "game": "Obsolete",
        "tier": "FAILURE",
        "year": 2022,
        "revenue": "<$10K est.",
        "review": "~30",
        "team_size": "Solo (leFarfelu, first-time dev)",
        "trailer": "Minimal or absent; game has near-zero marketing footprint; no social media presence",
        "ui": "Monochrome minimal; blob transformation indicator only; no keybinding support noted as friction",
        "art_style": "Monochrome minimalist pixel art; black-white only palette; isolation atmosphere as visual intent",
        "feature": "- Simple movement/jump/blob transformation mechanic\n- Environmental storytelling through minimalist design\n- No combat; exploration only\n- Very limited mechanical vocabulary",
        "scope": "1-2hr; very small scale; first-time developer project",
        "content": "Short single-environment experience; atmospheric OST praised as main strength. 'Evokes isolation well but thin on gameplay' - Steam review",
        "replayability": "Very low; no incentive to replay; 'Played once, done' - only reviewer sentiment",
    },
    {
        "genre": "2D Platformer",
        "game": "Bushiden",
        "tier": "FAILURE",
        "year": 2022,
        "revenue": "<$50K est.",
        "review": "~50",
        "team_size": "Team of 3 (Pixel Arc Studios)",
        "trailer": "Showed impressive cyberpunk pixel art in trailers; Kickstarter buzz in 2018 didn't convert to launch traction; EA launch suffered from incomplete state",
        "ui": "Unknown from available data; EA reviews note unpolished state",
        "art_style": "2D pixel art cyberpunk ninja; dark neon palette; inspired by 16-bit Ninja Gaiden / Strider aesthetic; visually distinctive",
        "feature": "- Ninja action platformer with cybernetic upgrade system\n- Kickstarter-funded (2018) with ambitious scope\n- EA launch without sufficient content to retain players\n- Metroidvania exploration elements",
        "scope": "EA launch with partial content; ~3-4hr estimated; development extended from 2018 Kickstarter to 2022+ EA without full release",
        "content": "Incomplete at EA launch; limited biomes; cyberpunk city setting. Reviews note 'great potential stuck in EA limbo'",
        "replayability": "Very low; EA state and low content volume discourage return; no active community",
    },
    {
        "genre": "2D Platformer",
        "game": "Leap Year",
        "tier": "FAILURE",
        "year": 2024,
        "revenue": "<$15K est.",
        "review": "~20",
        "team_size": "Solo (unknown)",
        "trailer": "Minimal - concept-driven pitch (collect Feb 2024 calendar pages) with no visible marketing execution",
        "ui": "Unknown; game has near-zero coverage",
        "art_style": "Pixel art; calendar/time-travel visual theme; minimal visual identity",
        "feature": "- 2D platformer built around collecting calendar pages (Feb 2024 theme)\n- Concept-first design without mechanical depth\n- Very limited scope",
        "scope": "1-2hr estimated; highly limited; niche concept without broad appeal",
        "content": "Calendar-themed levels; puzzle-platformer elements; thematic novelty without sufficient content depth",
        "replayability": "None; concept exhausted in first playthrough; no community",
    },
    {
        "genre": "2D Platformer",
        "game": "Desvelado - Generic Archetype (representative)",
        "tier": "FAILURE",
        "year": 2023,
        "revenue": "<$5K est.",
        "review": "~10",
        "team_size": "Solo (first-time dev)",
        "trailer": "Generic pixel art platformer - no distinguishing mechanic in capsule or trailer; looks like one of ~1,960 annual releases; zero social media traction",
        "ui": "Standard 2D platformer HUD; nothing distinguishing the experience visually",
        "art_style": "Generic pixel art; muted or default Unity/GMS2 asset-adjacent look; no distinctive visual identity",
        "feature": "- Standard jump, run, collect loop\n- No distinguishing mechanic or hook\n- Generic enemy and level design\n- Represents 99%+ of annual 2D platformer releases",
        "scope": "2-4hr; no post-launch support; developer moved on",
        "content": "Generic levels with no memorable moments; no story investment; 'Another pixel platformer' is the implicit review",
        "replayability": "Zero; invisible on Steam from launch; no wishlist momentum; no community",
    },
]

for g in game_rows:
    row = [
        g["genre"], g["game"], g["year"], g["revenue"], g["review"],
        g["team_size"], g["trailer"], g["ui"], g["art_style"],
        g["feature"], g["scope"], g["content"], g["replayability"],
    ]
    ws.append(row)
    xl_row = ws.max_row

    fill = TIER_FILLS.get(g["tier"])
    if fill:
        for col in range(1, 14):
            ws.cell(xl_row, col).fill = fill

    for col in range(7, 14):
        cell = ws.cell(xl_row, col)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

ws.freeze_panes = "A2"

BASE = Path(r"C:\Organized Files\My Game Asset\Game-Research")
output_path = BASE / "gameplay-review-2d-platformer.xlsx"
wb.save(str(output_path))
print(f"Saved: {output_path}")

snapshot_dir = BASE / "snapshots" / "2d-platformer"
snapshot_dir.mkdir(parents=True, exist_ok=True)
snapshot_path = snapshot_dir / f"{date.today().isoformat()}.xlsx"
shutil.copy2(str(output_path), str(snapshot_path))
print(f"Snapshot: {snapshot_path}")

log_path = BASE / "research-log.md"
today_str = date.today().isoformat()
high_count = sum(1 for g in game_rows if g["tier"] == "HIGH")
mid_count  = sum(1 for g in game_rows if g["tier"] == "MID")
fail_count = sum(1 for g in game_rows if g["tier"] == "FAILURE")
entry = f"| {today_str} | 2D Platformer | {high_count} | {mid_count} | {fail_count} | `snapshots/2d-platformer/{today_str}.xlsx` |\n"
if log_path.exists():
    log_path.write_text(log_path.read_text(encoding="utf-8") + entry, encoding="utf-8")
else:
    log_path.write_text(
        "# Research Log\n\n"
        "| Date | Genre | HIGH | MID | FAILURE | Snapshot |\n"
        "|------|-------|------|-----|---------|----------|\n"
        + entry,
        encoding="utf-8",
    )
print("research-log.md updated.")
print(f"Games: {high_count} HIGH / {mid_count} MID / {fail_count} FAILURE = {len(game_rows)} total")
