import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from pathlib import Path
import shutil
from datetime import date
import csv

HEADER_FILL = PatternFill('solid', fgColor='CFE2F3')
HIGH_FILL   = PatternFill('solid', fgColor='D9EAD3')
MID_FILL    = PatternFill('solid', fgColor='FFF2CC')
FAIL_FILL   = PatternFill('solid', fgColor='FFE6E6')
TIER_FILLS  = {'HIGH': HIGH_FILL, 'MID': MID_FILL, 'FAILURE': FAIL_FILL}

genre = '3D Metroidvania'
genre_slug = '3d-metroidvania'

game_rows = [
    # ---- HIGH TIER (7 games) ----
    {
        'genre': genre, 'tier': 'HIGH', 'game': 'Tunic', 'year': 2022,
        'revenue': '~$25M', 'review': '~17,000', 'team_size': 'Solo (Andrew Shouldice)',
        'trailer': 'Mystery teaser withheld gameplay for weeks; final 90s trailer: isometric fox in cryptic world, builds wonder without spoiling. One of the best-crafted indie reveals ever.',
        'ui': 'Alien-script in-game manual IS the UI; near-zero on-screen HUD; map fully diegetic; deliberate opacity as design feature.',
        'art_style': 'Low-poly isometric 3D; warm vibrant greens with ruins; cel-shaded palette; charming fox protagonist.',
        'feature': 'Isometric metroidvania; cryptic manual players decode; ability-gated secrets; combat; multiple endings; post-game challenge mode; world designed to be re-read.',
        'scope': '15-25hr main; post-game content; 3 major zones + secret areas; slow-revelation progression pacing.',
        'content': 'Core world + secret basement + final zone; manual collectibles; NG+ inverted mode; very dense with secrets. "Far more than expected after 20hrs" -- Steam review.',
        'replayability': 'Medium-high -- NG+ inverted mode, active speedrun community, manual-decoding replay. "Still finding things after 20hrs" -- Reddit r/TunicGame.',
    },
    {
        'genre': genre, 'tier': 'HIGH', 'game': 'Kena: Bridge of Spirits', 'year': 2021,
        'revenue': '~$25M', 'review': '~17,188', 'team_size': 'Ember Lab (~25 devs)',
        'trailer': 'Pixar-quality CG intro; 2min launch trailer shows combat, exploration, world traversal, Rot companions -- exceptional marketing for first-time studio.',
        'ui': 'Clean minimal third-person HUD; Rot companion count elegant; ability wheel uncluttered; zero information overload.',
        'art_style': 'Animated film quality 3D; vibrant lush forest with creeping dark corruption; spirit light particles; photo-realistic vegetation.',
        'feature': 'Third-person action-adventure; Rot companion spirits (assist combat + puzzles); ability-gated exploration; staff + bow combat; photo mode; accessible difficulty scaling.',
        'scope': '8-12hrs; 3 spirit guide zones + village hub; boss arena time trials; NG+ available.',
        'content': 'Main story + spirit guide challenges + collectible Rot hats + photo mode. "More polished than most AAA games" -- Steam.',
        'replayability': 'Medium -- NG+ hard mode; spirit guide time trials; Rot collectibles; growing speedrun community. Completionist tracker built in.',
    },
    {
        'genre': genre, 'tier': 'HIGH', 'game': "Another Crab's Treasure", 'year': 2024,
        'revenue': '~$15M', 'review': '~12,900', 'team_size': 'Aggro Crab (~6 devs)',
        'trailer': 'Contrast comedy: cute crab in brutal soulslike world; shell-swapping shown clearly; accessibility gun option prominently featured; excellent launch window visibility.',
        'ui': 'Clean for soulslike genre; shell HP system visual; accessibility gun option visible in menus; never overwhelming.',
        'art_style': 'Vibrant cartoon 3D ocean floor; trash-shell aesthetic; sun-dappled underwater lighting; personality-filled enemy design.',
        'feature': 'Soulslike combat (parry/dodge/attack); shell equipping system (30+ shells = armor); interconnected ocean map; ability gating via gear and movement abilities; accessibility gun mode.',
        'scope': '15-25hrs; 6+ ocean biomes; 30+ shells; multiple boss tiers; extensive optional content and secrets.',
        'content': 'Full metroidvania ocean map; hard optional bosses; gun accessibility mode brings new audiences; NG+ speedruns. "100+ hrs possible for completionists" -- Reddit r/AnotherCrabs.',
        'replayability': 'High -- shell variety creates build diversity; NG+ challenge mode; speedrun community; accessibility modes widen audience significantly.',
    },
    {
        'genre': genre, 'tier': 'HIGH', 'game': 'Supraland', 'year': 2019,
        'revenue': '$7-10M', 'review': '~14,645', 'team_size': 'Solo (David Muennich)',
        'trailer': '"Sandbox sandbox sandbox" concept shown immediately; toy-world aesthetic clear; Portal/Zelda/Metroid comparisons stated explicitly; humor from first frame.',
        'ui': 'Near-zero HUD; optional hint system; first-person clean view; no map clutter.',
        'art_style': 'First-person toy-box 3D; bright primary colors; sandbox world scaled to toy humans; charming cartoonish everyday objects as epic obstacles.',
        'feature': 'First-person metroidvania + Portal-style physics puzzles; exploration without hand-holding; combat; hidden passages; complete freedom to find own path.',
        'scope': '10-15hrs; one fully connected sandbox world; ~20 abilities/upgrades; secret areas everywhere; no fast travel initially forces world memory.',
        'content': 'Entirely connected world; ability unlocks make backtracking rewarding; DLC expansions available. "Found new things on 3rd playthrough" -- Reddit r/supraland.',
        'replayability': 'Medium -- high secret density rewards revisiting; Crash + Six Inches Under DLCs extend; speedrunning community; puzzle variety holds up on replay.',
    },
    {
        'genre': genre, 'tier': 'HIGH', 'game': 'Pseudoregalia', 'year': 2023,
        'revenue': '~$600K', 'review': '~10,000', 'team_size': 'Solo (rittzler)',
        'trailer': 'Pure 30s gameplay loop -- shows wall-jumps, backflips, air dashes; movement feel IS the entire pitch; $6 price shown prominently; immediate desire response.',
        'ui': 'Nearly invisible -- health as screen-border visual; no map cluttering screen; zero intrusion on movement experience.',
        'art_style': 'Lo-fi N64/PS1 nostalgic 3D; hazy dream-castle in soft purple and gold; textures intentionally degraded; Gothic castle atmosphere.',
        'feature': 'Movement-focused 3D metroidvania; wall jump; backflip; air dash; dream kick; non-linear castle; passive enemies (no combat required); pure movement exploration; $6.',
        'scope': '3-5hrs; castle + surrounding courtyards; ~10 movement abilities; zero padding; extremely tight and intentional design.',
        'content': 'Dense secret routes; movement tech depth far exceeds content length; speedrunners found shortcuts devs missed. "$6 for this is insane" -- Steam reviews consistently.',
        'replayability': 'Very high -- movement mastery curve; speedrunning extremely active (100+ leaderboard holders); community discovers new routes continuously; jam origin = tight design.',
    },
    {
        'genre': genre, 'tier': 'HIGH', 'game': 'Lunacid', 'year': 2022,
        'revenue': '$2.6M', 'review': '~7,648', 'team_size': 'Kira LLC (Jonathan Dean + small team)',
        'trailer': 'Dark atmospheric dungeon teaser; King\'s Field + Shadow Tower comparisons stated; moody OST; niche but perfectly community-targeted marketing.',
        'ui': 'Retro-styled first-person stat display; inventory screen; dark fantasy aesthetic matches game world exactly.',
        'art_style': 'First-person retro 3D dungeon; deep shadow pools; muted stone-and-torchlight palette; Shadow Tower / King\'s Field PS1-era inspiration.',
        'feature': 'First-person dungeon crawler; melee + ranged + magic combat; fully connected dungeon world; lore hidden everywhere; optional NPC quests; stat-gated progression.',
        'scope': '10-20hrs; 6+ dungeon regions; extensive item variety; multiple hidden passages; very long for $15 price.',
        'content': 'Multiple dungeon biomes; secret areas; extensive lore books; NPC sidequests; different endings. "Found something new after 40hrs" -- Steam.',
        'replayability': 'High for target audience -- stat builds vary combat feel; lore completionists; different weapon playstyles; retro fans return for atmosphere.',
    },
    {
        'genre': genre, 'tier': 'HIGH', 'game': 'Hob', 'year': 2017,
        'revenue': '~$1.5M', 'review': '~1,918', 'team_size': 'Runic Games (~20 devs, studio closed after)',
        'trailer': 'Gorgeous isometric world-transformation mechanic shown; robot protagonist; atmospheric music; zero UI in trailer -- tells you what game IS without words.',
        'ui': 'Zero UI -- completely HUD-free by design; all information conveyed through world state; bold design choice that pays off.',
        'art_style': 'Isometric 3D; vibrant greens + blues vs corrupted dark reds; mechanical robot aesthetic; world itself is the visual storytelling.',
        'feature': 'Isometric 3D metroidvania; world physically transforms as you progress; mechanical arm upgrades; no dialogue; no text; pure visual storytelling throughout.',
        'scope': '8-12hrs; 4 distinct areas; world physically changes with upgrades; 8 upgrade paths; Runic\'s final game before closure.',
        'content': 'Interconnected mechanical world that reshapes itself; hidden areas throughout; bittersweet visual narrative; studio\'s swan song. "World changing around you is unforgettable" -- Steam.',
        'replayability': 'Low -- primarily single emotional playthrough; studio closed post-launch; unique world transformation drives completionist first run.',
    },

    # ---- MID TIER (7 games) ----
    {
        'genre': genre, 'tier': 'MID', 'game': 'PowerSlave Exhumed', 'year': 2022,
        'revenue': '~$900K', 'review': '723', 'team_size': 'Nightdive Studios (~15 devs)',
        'trailer': 'Retro FPS revival pitch; Egyptian mythology aesthetic; classic gameplay shown; nostalgia-targeting fans of the 1996 original.',
        'ui': 'Classic FPS HUD; weapon selector; health/ammo display; authentic retro presentation.',
        'art_style': 'Modernized classic FPS 3D; Egyptian pyramid tombs; 90s aesthetic with modern lighting enhancements.',
        'feature': 'First-person shooter with full metroidvania backtracking; ability gating in Egyptian world; classic weapons + new movement abilities; Nightdive KEX engine enhancement.',
        'scope': '6-10hrs; 6 episodes; ability-required secrets throughout; full remaster with metroidvania structure intact.',
        'content': 'Full faithful remaster; all metroidvania secrets preserved; enhanced controls; extra accessibility modes added.',
        'replayability': 'Medium -- classic experience for FPS fans; speedrunning; new players discovering original game\'s unique metroidvania structure.',
    },
    {
        'genre': genre, 'tier': 'MID', 'game': 'The Last Case of Benedict Fox', 'year': 2023,
        'revenue': '~$680K', 'review': '~906', 'team_size': 'Plot Twist (~6 devs)',
        'trailer': 'Atmospheric Lovecraftian detective noir; memory-diving mechanic shown beautifully; dark oil-painting art front and center -- strong hook for art and narrative fans.',
        'ui': 'Some menu clutter from investigation system; journal mechanic integrated; manageable but busier than ideal.',
        'art_style': 'Hand-painted 2.5D-3D Gothic; rich dark oil-painting aesthetic; Lovecraftian creature design; one of the most visually striking indie games of 2023.',
        'feature': 'Metroidvania + detective mystery; dive into murder victims\' memories; demon companion abilities; puzzle solving; lore-heavy narrative; Definitive Edition adds content.',
        'scope': '8-12hrs; 6+ memory worlds; full murder mystery narrative; Definitive Edition improved reception and added story.',
        'content': 'Detective story across multiple memory worlds; collectible evidence; Lovecraftian horror atmosphere; Definitive Edition significantly improved the game.',
        'replayability': 'Low -- narrative-driven; completionist replay only; combat too simplistic for repeat runs.',
    },
    {
        'genre': genre, 'tier': 'MID', 'game': 'Coridden', 'year': 2022,
        'revenue': '~$600K', 'review': '402', 'team_size': 'Massive Miniteam (~6-10 devs)',
        'trailer': 'Shapeshifting mechanic shown front and center; dark dungeon tone; co-op gameplay prominently featured; action-forward presentation.',
        'ui': 'Dark fantasy RPG HUD; stat display; ability bar; competent but standard.',
        'art_style': 'Dark fantasy 3D; gritty dungeon environments; each monster form has distinct visual identity; competent but not standout.',
        'feature': 'Shapeshifting RPG -- absorb monster DNA; 3D hack-and-slash; co-op multiplayer; light metroidvania exploration; ability gating per beast form.',
        'scope': '10-15hrs; dungeon + city hub; multiple beast forms; limited endgame content criticized.',
        'content': 'Main quest + side content; beast collection system; solo or co-op; needs more boss variety and polish.',
        'replayability': 'Low-medium -- co-op adds replay value; beast hunting encourages exploration; limited endgame hurts long-term retention.',
    },
    {
        'genre': genre, 'tier': 'MID', 'game': 'After Us', 'year': 2023,
        'revenue': '~$700K', 'review': '308', 'team_size': 'Piccolo Studio (~10 devs)',
        'trailer': 'Environmental art showcased beautifully; surrealist post-apocalyptic vision; emotional tone with animal spirits; no combat shown -- art-house positioning.',
        'ui': 'Extremely minimal -- spirit aura acts as health indicator; essentially invisible UI; perfect fit for artistic vision.',
        'art_style': 'Surrealist 3D; crumbling post-apocalyptic earth with organic spirit elements; ethereal particle effects; stunning environmental design throughout.',
        'feature': 'Non-combat 3D platformer; spirit animal collection; ability-gated progression; environmental storytelling; emotional narrative about earth\'s lost wildlife.',
        'scope': '5-8hrs; 6+ distinct biome areas; focused story; zero padding; each level has unique visual theme.',
        'content': 'Spirit animals to collect in each zone; environmental lore throughout; focused narrative about extinction; emotional payoff ending.',
        'replayability': 'Very low -- short; no combat; primarily single-playthrough art experience; spirit collectible replay only.',
    },
    {
        'genre': genre, 'tier': 'MID', 'game': 'Haydee 3', 'year': 2025,
        'revenue': '~$400K', 'review': '301', 'team_size': 'Solo (MrDrNose)',
        'trailer': 'Android protagonist highlighted; industrial maze environment; hardcore difficulty signaled immediately; distinctive character design is the franchise hook.',
        'ui': 'Functional but minimal; inventory screen; health display; map available but deliberately limited.',
        'art_style': 'Stylized 3D; android protagonist with signature design; color-coded industrial sections; functional aesthetics with series identity.',
        'feature': 'Hardcore 3D metroidvania puzzle-platformer; third-person shooter sections; pure maze navigation; no hand-holding; 160+ rooms; series tradition of uncompromising design.',
        'scope': '10-20hrs for skilled players; 6 themed sections; 160 rooms; very high difficulty maintained throughout.',
        'content': '6 themed sections; 160 rooms; enemy AI; hidden items and paths; very punishing difficulty -- rewards mastery and route memorization.',
        'replayability': 'High for dedicated fans -- route optimization; difficulty mastery curve; solo dev ongoing support; dedicated series community.',
    },
    {
        'genre': genre, 'tier': 'MID', 'game': 'Supraland Crash', 'year': 2020,
        'revenue': '~$350K', 'review': '579', 'team_size': 'Solo (David Muennich, DLC/expansion)',
        'trailer': 'Short; directly addresses existing Supraland fans; shows new chaotic area style; standalone experience pitch.',
        'ui': 'Identical to Supraland base game.',
        'art_style': 'Same toy-box aesthetic as Supraland; more experimental and chaotic visual feel; smaller scale.',
        'feature': 'First-person metroidvania puzzles; standalone from main game; more experimental design; shorter and more concentrated.',
        'scope': '5-8hrs; smaller connected area; ~10 puzzles/abilities; chaotic design philosophy; clear expansion intent.',
        'content': 'New area; different puzzle philosophy; more experimental than base game; fans appreciated the experimentation.',
        'replayability': 'Low -- supplement for Supraland fans only; no NG+; puzzle-once design by nature.',
    },
    {
        'genre': genre, 'tier': 'MID', 'game': 'Recompile', 'year': 2021,
        'revenue': '~$213K', 'review': '142', 'team_size': 'Phigames (~4 devs)',
        'trailer': 'Stylized cyber aesthetic; glitch visual effects; abstract 3D cyberspace; narrative AI ethics hook -- artistically impressive but undersells gameplay.',
        'ui': 'Styled to match cyber aesthetic; some readability issues; HUD fits thematic world even if slightly hard to parse.',
        'art_style': 'Abstract 3D cyberspace; neon purple on dark background; glitch visual effects; data architecture rendered as physical game space.',
        'feature': '3D metroidvania with hacking mechanic; air-dash movement; ethical choice system determines endings; combat + puzzle + platforming; narrative-led throughout.',
        'scope': '5-8hrs; connected virtual world; branching moral endings; small but focused; published by Humble Games.',
        'content': 'Multiple endings based on AI ethics choices; cyber lore; ability-gated world design; air dash central to movement. "Concept stronger than execution" -- common review sentiment.',
        'replayability': 'Medium -- multiple endings incentivize replay; rough combat feel is the barrier; short enough that replay is feasible for committed fans.',
    },

    # ---- FAILURE TIER (4 verified -- niche genre limits pool) ----
    {
        'genre': genre, 'tier': 'FAILURE', 'game': "Jett: The Far Shore", 'year': 2021,
        'revenue': '~$126K', 'review': '56', 'team_size': 'Superbrothers + Pine Scented (~8 devs)',
        'trailer': 'Minimalist space odyssey; artistic and atmospheric but almost no gameplay shown; extremely slow pacing alienated general audience; launched Epic-exclusive.',
        'ui': 'Minimalist; fits sparse aesthetic but feels empty rather than purposeful.',
        'art_style': 'Minimalist 3D; abstract alien landscapes; watercolor-influenced environments; Superbrothers signature style in 3D with mixed results.',
        'feature': 'Space exploration via scout vehicle; alien planet discovery; ability-gated biome access; narrative-heavy; no combat; meditation pace throughout.',
        'scope': '5-8hrs; multiple alien planets; heavily narrative-driven; limited player agency criticized widely.',
        'content': 'Alien worlds with subtle lore; companion characters; narrative climax; very small scope; Epic exclusive through 2021-2022 severely limited Steam reach.',
        'replayability': 'Very low -- pure narrative; minimal exploration freedom; Epic exclusivity damaged discovery; now largely forgotten despite Superbrothers pedigree.',
    },
    {
        'genre': genre, 'tier': 'FAILURE', 'game': 'Tamarin', 'year': 2020,
        'revenue': '~$142K', 'review': '76', 'team_size': 'Chameleon Games (~15 devs, ex-Rare veterans)',
        'trailer': 'Colorful monkey in Northern nature; raises Banjo-Kazooie comparisons the dev team helped create; shows exploration and shooting; creates expectations gameplay cannot meet.',
        'ui': 'Standard 3D platformer HUD; nothing remarkable; competent but forgettable.',
        'art_style': '3D colorful Northern nature environment; monkey protagonist; technically adequate but lacks personality of its spiritual inspirations.',
        'feature': '3D platformer + third-person shooter hybrid; collect-a-thon elements; metroidvania world exploration; shooting sections severely interrupt platforming flow.',
        'scope': '8-12hrs; nature wilderness areas; collectibles throughout; disappointing execution of a compelling concept from talented ex-Rare team.',
        'content': 'Multiple nature areas; collectibles; platforming/shooting mix; unfavorable Banjo-Kazooie comparison dominated all reception.',
        'replayability': 'Very low -- mixed execution prevents replay; ex-Rare legacy was double-edged sword; game promised more than it delivered.',
    },
    {
        'genre': genre, 'tier': 'FAILURE', 'game': 'Testament: The Order of High Human', 'year': 2023,
        'revenue': '~$93K', 'review': '83', 'team_size': 'Fairyship Games (~4 devs)',
        'trailer': 'First-person swordplay in fantasy dungeon shown; ambitious concept for a small team; execution concerns visible in trailer before launch.',
        'ui': 'RPG elements overlay; some clutter; small budget constraints visible in UI design.',
        'art_style': 'Fantasy first-person 3D; decent dungeon corridor environments; competent lighting; budget constraints limit visual ambition.',
        'feature': 'First-person swordplay + metroidvania elements + RPG stats; puzzle-solving; ability gating; unfortunately rough combat feel undermines an interesting concept.',
        'scope': '8-12hrs; castle and dungeon exploration; RPG stat progression; too rough overall to be satisfying.',
        'content': 'Full narrative; dungeon exploration; ability unlocking; consensus: interesting concept hampered by low polish. Only 42% positive reviews.',
        'replayability': 'Very low -- mostly negative reception; core combat feel issues prevent any replay motivation.',
    },
    {
        'genre': genre, 'tier': 'FAILURE', 'game': 'Hellscreen', 'year': 2026,
        'revenue': '~$36K', 'review': '44', 'team_size': 'Mixtape Games UK (Solo/Duo)',
        'trailer': 'Rear-view mirror shooting mechanic shown prominently; doom-inspired corridors; retro aesthetic communicated clearly; extremely niche target audience.',
        'ui': 'Classic retro FPS HUD; functional and thematically consistent.',
        'art_style': 'Retro doom-style FPS corridors; dark minimal color palette; deliberately low-budget aesthetic.',
        'feature': 'FPS + rearview mirror mechanic (shoot what is behind you); metroidvania hub with ability unlocks; Original concept; shipped shorter than planned EA roadmap.',
        'scope': 'Very short -- launched 3 episodes vs planned 5; ~3-4hrs total; scope reduced from EA vision.',
        'content': '3 episodes (planned 5 cut); metroidvania hub connecting levels; rearview gimmick is genuinely novel but too short to build audience.',
        'replayability': 'Low -- too short; gimmick wears thin quickly; no community formed around it.',
    },
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Genre Deep Dive'

headers = ['Genre','Game','Year','Revenue','Review','Team size','Trailer','UI','Art style','Feature','Scope','Content','Replayability']
ws.append(headers)
for cell in ws[1]:
    cell.fill = HEADER_FILL
    cell.font = Font(bold=True)

col_widths = {'A':20,'B':28,'C':8,'D':16,'E':10,'F':22,'G':36,'H':28,'I':24,'J':52,'K':38,'L':48,'M':48}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

for g in game_rows:
    row = [g['genre'],g['game'],g['year'],g['revenue'],g['review'],g['team_size'],g['trailer'],g['ui'],g['art_style'],g['feature'],g['scope'],g['content'],g['replayability']]
    ws.append(row)
    xl_row = ws.max_row
    fill = TIER_FILLS.get(g['tier'])
    if fill:
        for col in range(1, 14):
            ws.cell(xl_row, col).fill = fill
    for col in range(7, 14):
        cell = ws.cell(xl_row, col)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

ws.freeze_panes = 'A2'

today = date.today()
output_path = r'C:\Organized Files\My Game Asset\Game Market\Game-Research\exports\gameplay-review-3d-metroidvania.xlsx'
wb.save(output_path)
print(f'Saved: {output_path}')

snapshot_dir = Path(r'C:\Organized Files\My Game Asset\Game Market\Game-Research\snapshots\3d-metroidvania')
snapshot_dir.mkdir(parents=True, exist_ok=True)
snapshot_path = snapshot_dir / f'{today.isoformat()}.xlsx'
shutil.copy2(output_path, snapshot_path)
print(f'Snapshot: {snapshot_path}')

log_path = Path(r'C:\Organized Files\My Game Asset\Game Market\Game-Research\research-log.md')
today_str = today.isoformat()
high_count = sum(1 for g in game_rows if g['tier'] == 'HIGH')
mid_count  = sum(1 for g in game_rows if g['tier'] == 'MID')
fail_count = sum(1 for g in game_rows if g['tier'] == 'FAILURE')
entry = f'| {today_str} | 3D Metroidvania | {high_count} | {mid_count} | {fail_count} | snapshots/3d-metroidvania/{today_str}.xlsx |\n'
if log_path.exists():
    log_path.write_text(log_path.read_text(encoding='utf-8') + entry, encoding='utf-8')
else:
    log_path.write_text(
        '# Research Log\n\n| Date | Genre | HIGH | MID | FAILURE | Snapshot |\n|------|-------|------|-----|---------|----------|\n' + entry,
        encoding='utf-8',
    )
print('research-log.md updated.')
print(f'Total games: {len(game_rows)} ({high_count} HIGH / {mid_count} MID / {fail_count} FAILURE)')
