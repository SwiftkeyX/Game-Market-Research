import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import date
import shutil
from pathlib import Path

HEADER_FILL = PatternFill("solid", fgColor="CFE2F3")
HIGH_FILL   = PatternFill("solid", fgColor="D9EAD3")
MID_FILL    = PatternFill("solid", fgColor="FFF2CC")
FAIL_FILL   = PatternFill("solid", fgColor="FFE6E6")
TIER_FILLS  = {"HIGH": HIGH_FILL, "MID": MID_FILL, "FAILURE": FAIL_FILL}

game_rows = [
  # ── HIGH ──────────────────────────────────────────────────────────────────
  {
    "tier": "HIGH", "genre": "Roguelike", "game": "Hades", "year": 2020,
    "revenue": "~$85M+", "review": "~303,000",
    "team_size": "~20 (Supergiant Games)",
    "trailer": "Cinematic opening, combat showcased in first 10s, Zagreus voiced, punchy editing, ~90s. Clear core loop from opening shot.",
    "ui": "Clean isometric HUD, health/gold top-right, boon menu elegant, no clutter mid-combat. Readable at all times.",
    "art_style": "Hand-painted 2D, vibrant Greek mythology aesthetic, high contrast, rich warm color palette",
    "feature": "Action roguelite + narrative-per-run unlocking; boon system (6 gods, 40+ boons each); 6 weapons with aspects; Heat difficulty scaling; NPC gift system",
    "scope": "6-10hr to first clear; 50+ runs to see all story; 6 weapons; EA Dec 2018 -> full Sept 2020",
    "content": "6 weapons, 40+ enemy types, 12 boss variants, voiced NPC relationship arcs. 'Story keeps unfolding after 100 runs' -- Steam",
    "replayability": "Extremely high -- narrative grows each run, weapon aspects, Heat system, gift unlocks. 'Still finding new dialogue at 200hrs' -- r/HadesTheGame",
  },
  {
    "tier": "HIGH", "genre": "Roguelike", "game": "Vampire Survivors", "year": 2022,
    "revenue": "~$30M+", "review": "~262,000",
    "team_size": "Solo (poncle / Luca Galante); small team post-launch",
    "trailer": "15s auto-gameplay clip, no voiceover, weapon evolution shown visually, $3 price on screen. Sold concept in under 20s.",
    "ui": "Minimal HUD, large visible XP gems, big readable pickup text, 6 weapon slots at screen edge. Intentionally lo-fi.",
    "art_style": "Pixel art, chunky sprites, dark gothic palette, exaggerated scale enemies",
    "feature": "Auto-shooter + weapon evolution combos; 30+ weapons with 50+ evolutions; 30+ characters; 6+ stages; DLC packs; secret unlocks",
    "scope": "30min per run; 30+ weapons; 30+ characters; 6 stages + DLC; EA Oct 2021 -> full release",
    "content": "30 characters, 30 weapons, evolution combos, hidden achievements, DLC stages. 'Still finding secrets after 200hrs' -- Steam",
    "replayability": "Very high -- character variety, build synergy, hidden unlocks, achievement hunting, DLC. 'Go-to comfort game, always one more run' -- Steam",
  },
  {
    "tier": "HIGH", "genre": "Roguelike", "game": "Balatro", "year": 2024,
    "revenue": "~$100M+", "review": "~191,000",
    "team_size": "Solo (LocalThunk)",
    "trailer": "Hypnotic loop: poker hand -> joker combo -> score explosion; no voiceover; synth music; 90s; compulsive pacing. Instantly conveys fantasy.",
    "ui": "Clean poker table aesthetic, neon accents on dark background, joker grid always visible, score multiplier prominent",
    "art_style": "Stylized poker card aesthetic, neon dark palette, bold numerics, casino visual language",
    "feature": "Poker hand scoring + 150 joker combos + consumables + 8 blinds; 8 decks; tarot/planet/spectral cards; 8 difficulty stakes; daily challenge",
    "scope": "3-6hr per run to beat base game; 150+ jokers; 8 decks; 8 stakes; launched Feb 2024 full release",
    "content": "150 jokers, 8 decks, 8 stakes, planet/tarot cards, challenges, daily mode. 'Every run feels fundamentally different' -- Steam",
    "replayability": "Extremely high -- joker combinations near-infinite; daily challenges; unlockable build paths. 'Most addictive game in years' -- r/balatro",
  },
  {
    "tier": "HIGH", "genre": "Roguelike", "game": "Hades II", "year": 2025,
    "revenue": "~$50M+", "review": "~119,000",
    "team_size": "~25 (Supergiant Games)",
    "trailer": "Cinematic intro with Melinoe vs Chronos framing; combat showcase at 20s; strong music; story-forward; ~2min. Fans of original immediately engaged.",
    "ui": "Similar to Hades 1 but cleaner; improved resource tracking sidebar; elemental indicators added; moon resource meter new",
    "art_style": "Hand-painted 2D, darker/more muted than Hades 1, moon goddess aesthetic, Underworld depth and layering",
    "feature": "New protagonist Melinoe; moonstone/arcana card progression; 6 weapons; fishing; herb crafting; 30+ gods with boons; deeper build synergy than Hades 1",
    "scope": "8-15hr to first clear; 6 weapons; deeper crafting; EA May 2024 -> full release Sept 2025",
    "content": "6 weapons, extensive NPC stories, 30+ gods, boss roster deeper than Hades 1. 'Even better than the original' -- Steam",
    "replayability": "Very high -- builds more complex, narrative evolves per run, multiple weapon forms. 'I thought nothing could top Hades 1, I was wrong' -- r/HadesTheGame",
  },
  {
    "tier": "HIGH", "genre": "Roguelike", "game": "Brotato", "year": 2023,
    "revenue": "~$10M+", "review": "~106,000",
    "team_size": "~2 (Blobfish, French studio)",
    "trailer": "45s fast-cut gameplay montage; upbeat electronic music; potato character visible immediately; weapon icons shown clearly. Pacing matches gameplay loop.",
    "ui": "Minimal in-run HUD; 6 weapon slots at screen edge; item pickup pop-ups clean; shop UI grid-based and scannable",
    "art_style": "Pixel art, cartoonish, bright primary colors, chunky potato character designs",
    "feature": "6-weapon loadout + item/stat synergies; 46 characters with unique starting stats/weapons; 5 difficulty waves per run; 60+ weapons; endless escalation",
    "scope": "10-30min per run; 46 characters; 60+ weapons; 5 difficulty modes; EA Aug 2022 -> full release",
    "content": "46 characters, 60+ weapons, hundreds of items, multiple difficulties. 'Every character is a completely different game' -- Steam",
    "replayability": "High -- character variety, build synergy, challenge modes, item unlocks. 'Fast enough that one more run is always possible' -- r/SurvivorsLike",
  },
  {
    "tier": "HIGH", "genre": "Roguelike", "game": "Halls of Torment", "year": 2023,
    "revenue": "~$1.5M", "review": "~31,300",
    "team_size": "~3-4 (Chasing Carrots, German studio)",
    "trailer": "Shows progression board, fast combat clips, Diablo-esque atmosphere; 60s; dark music. Nostalgic hook for ARPG fans lands immediately.",
    "ui": "Clean ARPG style; ability slots bottom-center; retro stat panel; Diablo 1-inspired interface",
    "art_style": "Low-poly/pixel hybrid, dark dungeon atmosphere, Diablo 1 retro aesthetic, intentional lo-fi",
    "feature": "Vampire survivor-like + Diablo-esque character classes + persistent skill board; 8 characters; ability upgrade tree; daily challenges",
    "scope": "20-30min per run; 8 characters; 5 stages; persistent upgrade board; EA June 2023 -> full release",
    "content": "8 characters, 5 stages, persistent upgrade system, daily challenges, item drops. 'Scratches my Diablo itch perfectly' -- Steam",
    "replayability": "High -- character variety, daily challenges, persistent progress, build optimization. 'My daily driver since launch' -- r/roguelikes",
  },
  {
    "tier": "HIGH", "genre": "Roguelike", "game": "Nubby's Number Factory", "year": 2025,
    "revenue": "~$2M+", "review": "~15,900",
    "team_size": "~3-4 (Everybody House Games)",
    "trailer": "Charming 30s clip, number manipulation gameplay shown in first 5s, quirky bouncy music, colorful factory aesthetic. Clear concept, surprising depth signal.",
    "ui": "Clean number grid, readable numeric displays, factory upgrade panel well-organized",
    "art_style": "Colorful cartoonish, factory/industrial aesthetic, playful numbered character designs",
    "feature": "Arithmetic/number puzzle mechanics + roguelike run structure; factory-building phase + run phase; unique chain number manipulation system",
    "scope": "2-4hr per run; multiple factory modes; upgrade paths; launched 2025 full release",
    "content": "Multiple factory stages, upgrade trees, combinatorial number systems. 'Surprisingly deep for a number game' -- Steam",
    "replayability": "Moderate-high -- build variety around number combos, different factory configurations. 'Every run plays differently' -- Steam",
  },

  # ── MID ───────────────────────────────────────────────────────────────────
  {
    "tier": "MID", "genre": "Roguelike", "game": "Rift Wizard 2", "year": 2023,
    "revenue": "~$400K", "review": "~728",
    "team_size": "Solo (Grifflan / Tom Hermans)",
    "trailer": "Text-focused, spellcasting grid shown, minimalist presentation. Functional rather than flashy -- niche appeal clearly signaled.",
    "ui": "Dense ASCII-inspired pixel UI; spell selection fills screen; designed for keyboard-only play; info-dense",
    "art_style": "Pixel art, dark fantasy palette, minimal animation, traditional roguelike aesthetics",
    "feature": "Turn-based grid spells + modifiers + branching runs; 200+ spells; permadeath; wizard build synergies; realm-jumping mechanic",
    "scope": "1-2hr per run; 200+ spells; procedural levels; short but dense decision-making per floor",
    "content": "200+ spells, enemy variety, permadeath pressure. 'Deepest spell-synergy system in any roguelike' -- r/roguelikes",
    "replayability": "High for core audience -- enormous spell variety, daily seeds. 'Still finding combos after 50 runs' -- Steam. Niche appeal caps reach.",
  },
  {
    "tier": "MID", "genre": "Roguelike", "game": "Go Mecha Ball", "year": 2024,
    "revenue": "~$413K", "review": "~367",
    "team_size": "~3-4 (VILE MONARCH)",
    "trailer": "Pinball-meets-shooter momentum gameplay shown clearly; neon aesthetic; dynamic movement showcased; ~60s. Concept is unique and visible.",
    "ui": "Clean arcade-style HUD; damage numbers visible; upgrade menu readable; health bar integrated into arena",
    "art_style": "Neon/cyberpunk 2D, cartoon with strong particle effects, vibrant color contrast",
    "feature": "Pinball physics momentum + twin-stick shooting + roguelike upgrades; 4 mechs; ball-form movement; upgrade combos",
    "scope": "1-2hr per run; 4 mechs; multiple zones; upgrade paths and passive combos",
    "content": "4 mechs, upgrade paths, boss fights. 'Criminally underrated momentum system' -- Vice. Content thinner than top-tier.",
    "replayability": "Moderate -- mech variety and upgrade paths, but shorter content set limits long-term play. 'Fun in short bursts' -- Steam",
  },
  {
    "tier": "MID", "genre": "Roguelike", "game": "Mortal Glory 2", "year": 2024,
    "revenue": "~$152K", "review": "~289",
    "team_size": "Solo (Windy Games)",
    "trailer": "Text-heavy tactical overview; gladiator arena shown; retro pixel aesthetic; functional presentation. Communicates niche appeal accurately.",
    "ui": "Clean grid-based tactical display; unit stat bars clear; action menu simple and legible",
    "art_style": "Pixel art, Roman/fantasy arena aesthetic, limited animation, functional design choices",
    "feature": "Tactical turn-based gladiator combat + roster management + roguelike run structure; permadeath; 50+ gladiator types; build specialization",
    "scope": "20-40min per run; 50+ gladiators; multiple arenas; permadeath roster; solo dev content volume",
    "content": "Multiple arenas, gladiator types, abilities. 'Excellent for tactical fans -- no bloat, just combat' -- r/roguelikes",
    "replayability": "Moderate-high -- roster variety and tactical challenge. 'Underappreciated gem' -- Steam. Solo dev limits total content volume.",
  },
  {
    "tier": "MID", "genre": "Roguelike", "game": "Rogue Hex", "year": 2025,
    "revenue": "~$247K", "review": "~253",
    "team_size": "Small team",
    "trailer": "Hex-grid combat shown clearly; strategy focus; functional but minimal presentation. Communicates genre to right audience.",
    "ui": "Hex grid with unit stat overlay; clean tile-based interface; clear visual hierarchy of tiles and units",
    "art_style": "Clean 2D hex grid, fantasy color palette, functional unit designs",
    "feature": "Hex-based turn-based strategy + roguelike run progression; tile control mechanics; unit abilities; map branching",
    "scope": "Multiple runs; branching hex maps; unit builds and ability synergies",
    "content": "Hex maps, enemy types, unit builds. 'Solid strategy roguelike for hex game fans' -- Steam",
    "replayability": "Moderate -- hex combat variety and build paths. Limited crossover appeal outside strategy fans.",
  },
  {
    "tier": "MID", "genre": "Roguelike", "game": "Katanaut", "year": 2025,
    "revenue": "~$50K", "review": "~175",
    "team_size": "Solo (Eugene / Voidmaw)",
    "trailer": "Action platformer roguelite gameplay shown; space/zombie theme; solid production. Buried by competing launches at release.",
    "ui": "Clean action platformer HUD; ability slots visible; health and currency readable",
    "art_style": "2D hand-drawn metroidvania-inspired aesthetic, sci-fi palette, space zombie theme",
    "feature": "Metroidvania-inspired action roguelite; sword combat + abilities; procedural levels; run-based progression",
    "scope": "1-2hr per run; procedural levels; multiple weapon types; 4+ years dev time",
    "content": "Solid action roguelite content -- game quality not the issue. 'Great game that nobody bought' -- dev postmortem. 87% positive reviews.",
    "replayability": "Moderate -- action variety and builds, solid loop. Timing disaster: launched between Silksong and Hades 2 with 52k wishlists but only 11 peak players.",
  },
  {
    "tier": "MID", "genre": "Roguelike", "game": "Iron Mandate", "year": 2024,
    "revenue": "$0 (made free)", "review": "~171",
    "team_size": "Solo (Gibson / Earthling Entertainment)",
    "trailer": "Shows mechanical sandworm side-scroller; steampunk aesthetic; functional but uninspired presentation",
    "ui": "Functional turret placement interface; somewhat cluttered panel; terrain visibility issues flagged by players",
    "art_style": "Pixel art, steampunk/mechanical, dark underground palette, low-detail terrain textures",
    "feature": "Side-scrolling roguelike + turret placement + mechanical sandworm locomotion; unique vehicle-based movement concept",
    "scope": "Short runs; limited level variety; EA period ended as free release; no further updates 23+ months",
    "content": "Sandworm mechanics, turret variety, limited enemy types. Developer cited 'lack of variety' as fatal design flaw.",
    "replayability": "Low -- wonky controls cited as primary drop-off trigger. Developer own postmortem: 'fails in fundamental ways'. Made free rather than continue.",
  },
  {
    "tier": "MID", "genre": "Roguelike", "game": "Deadzone: Rogue", "year": 2025,
    "revenue": "~$150K (est.)", "review": "~200 (est.)",
    "team_size": "Small team",
    "trailer": "Action roguelike gameplay shown; sci-fi military shooter aesthetic; competent functional presentation",
    "ui": "Standard action roguelite HUD; ability panel; health/resource bars; genre-conventional",
    "art_style": "2D sci-fi shooter aesthetic, military/tactical palette, solid execution",
    "feature": "Action roguelite + shooter mechanics; run-based loadout building; character variety",
    "scope": "Standard action roguelite run length; multiple characters and weapon loadouts",
    "content": "Multiple characters, weapon builds, run variety. Solid execution in a very crowded space.",
    "replayability": "Moderate -- genre-standard replay loop competing with stronger established titles",
  },

  # ── FAILURE ───────────────────────────────────────────────────────────────
  {
    "tier": "FAILURE", "genre": "Roguelike", "game": "Bramble Royale", "year": 2025,
    "revenue": "~$15K", "review": "~20",
    "team_size": "Solo (Slothwerks / Matthew Rader)",
    "trailer": "Deck-builder with team-based plant/nature aesthetic; competent but genre-standard. No distinct hook to separate from Balatro noise.",
    "ui": "Card-based grid interface; deck management familiar to genre veterans; nothing stands out visually",
    "art_style": "Cartoon fantasy, plant/nature aesthetic, warm palette, approachable",
    "feature": "Team-based roguelike deckbuilder; card synergies; nature/plant theme; upgrade paths. Solid execution of genre conventions.",
    "scope": "Standard deckbuilder run length; multiple characters; card variety",
    "content": "Good card content for genre fans. Problem was deckbuilder fatigue perception, not quality. Only 703 units sold in first week.",
    "replayability": "Moderate -- card variety and team builds. 'A great game that nobody found' -- developer postmortem. Grossed $8K in week 1.",
  },
  {
    "tier": "FAILURE", "genre": "Roguelike", "game": "Hardest", "year": 2025,
    "revenue": "~$0 (F2P)", "review": "31",
    "team_size": "Solo dev",
    "trailer": "Minimal functional presentation; basic card roguelike concept. No identity or hook visible.",
    "ui": "Basic card interface; functional but unpolished; no visual identity",
    "art_style": "Minimal/abstract, very limited visual identity, undifferentiated",
    "feature": "Card-based roguelike; basic combat mechanics; limited feature set; F2P model",
    "scope": "Short runs; limited content depth; F2P with no depth driver",
    "content": "Basic card variety; insufficient content for retention. 54% positive -- mixed reception from sparse audience. Being deleted from Steam Jan 2026.",
    "replayability": "Very low -- thin content, unpolished execution. Mixed rating reflects core design issues, not just visibility.",
  },
  {
    "tier": "FAILURE", "genre": "Roguelike", "game": "Enspell", "year": 2024,
    "revenue": "~$20K (est.)", "review": "~50 (est.)",
    "team_size": "Small team (~2-3)",
    "trailer": "Spell-based roguelike deckbuilder; EA presentation; promising concept visible but execution thin",
    "ui": "EA-quality interface; spell card layout; needs significant polish",
    "art_style": "Fantasy pixel/2D, spell aesthetic, unfinished visual identity at EA stage",
    "feature": "Spell/magic-themed deckbuilder roguelike; combo system; EA content set with roadmap promises",
    "scope": "EA-limited scope; development discontinued mid-roadmap; leaving Steam",
    "content": "Partial EA content; roadmap features never delivered. 'EA abandoned mid-development' -- Steam reviews. Leaving Steam, no future updates.",
    "replayability": "Low -- incomplete content, no updates 2+ years. Development ceased due to internal circumstances.",
  },
  {
    "tier": "FAILURE", "genre": "Roguelike", "game": "Realm of Ink", "year": 2024,
    "revenue": "~$10K (est.)", "review": "~50 (est.)",
    "team_size": "Small indie team",
    "trailer": "Ink/brush art aesthetic roguelike; distinctive visual hook; EA launch with strong art identity",
    "ui": "Stylized ink-themed interface; visual identity strong; gameplay thin at time of delisting",
    "art_style": "Ink brush/traditional Chinese painting aesthetic; distinctive black-ink visual style; genuinely unique",
    "feature": "Action roguelite with ink/brush power system; procedural runs; art-forward design",
    "scope": "EA-limited; delisted from Steam before reaching planned scope",
    "content": "Unique art style but thin gameplay content at time of delisting. Developer claims development continues but Steam page gone.",
    "replayability": "Unestablished -- delisted before audience could form. 'Great concept, not enough game' -- early players",
  },
  {
    "tier": "FAILURE", "genre": "Roguelike", "game": "Void War", "year": 2025,
    "revenue": "~$5K (est.)", "review": "~25 (est.)",
    "team_size": "Solo dev",
    "trailer": "FTL-inspired space roguelite; Warhammer 40K aesthetic; functional basic presentation",
    "ui": "FTL-inspired strategic map + combat interface; readable but heavily derivative of source material",
    "art_style": "2D sci-fi space, Warhammer 40K-inspired, derivative of GW IP visual language",
    "feature": "FTL-style space roguelite + Warhammer 40K theme; ship management; tactical combat; run-based",
    "scope": "Standard FTL-like run length; ship upgrade system; development cut short by DMCA",
    "content": "Solid FTL-like loop but legally compromised. Pulled from Steam via DMCA claim from Games Workshop before building any audience.",
    "replayability": "Irrelevant -- game removed by IP claim. FTL meets Warhammer is compelling concept, legally unsustainable without IP license.",
  },
  {
    "tier": "FAILURE", "genre": "Roguelike", "game": "Voids Vigil", "year": 2024,
    "revenue": "~$10K (est.)", "review": "~30 (est.)",
    "team_size": "Solo dev",
    "trailer": "Auto-shooter gameplay shown; described as lean Vampire Survivors experience; minimal production quality",
    "ui": "Simple auto-battle HUD; minimal visual complexity; nothing distinguishing",
    "art_style": "Pixel art, dark sci-fi palette, minimal character designs, low visual identity",
    "feature": "Auto-shooter roguelike; weapon pickups; wave survival; positioned as more concise Vampire Survivors alternative",
    "scope": "Short runs; limited weapon variety; minimal content versus established genre leaders",
    "content": "Derivative of Vampire Survivors with less content. 'Lean' conceded by dev, read as 'thin' to buyers with better options.",
    "replayability": "Low -- outcompeted by VS (cheap), Brotato (cheap), and numerous free demos. No differentiating hook.",
  },
  {
    "tier": "FAILURE", "genre": "Roguelike", "game": "AI Roguelite", "year": 2023,
    "revenue": "~$5K (est.)", "review": "~30 (est.)",
    "team_size": "Solo dev",
    "trailer": "AI-generated assets and concept; generic roguelike dungeon crawler presentation; no distinct style",
    "ui": "Basic dungeon crawler interface; unpolished UI; AI-generated art creates inconsistent visual identity",
    "art_style": "AI-generated art assets, inconsistent visual style, generic dungeon palette",
    "feature": "Generic dungeon roguelike; basic combat; AI-generated content; limited mechanical depth",
    "scope": "Short runs; limited variety; stuck in EA with game-breaking bugs; no working new game option for some players",
    "content": "Thin content, broken features, AI-generated assets. 'Cannot start new game anymore' -- Steam bug report. EA abandoned.",
    "replayability": "Very low -- bugs prevent play, AI art creates no identity, no community formed. Cautionary tale of asset-flip roguelikes.",
  },
]

# ── Excel ──────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Genre Deep Dive"

headers = ["Genre","Game","Year","Revenue","Review","Team size","Trailer","UI","Art style","Feature","Scope","Content","Replayability"]
ws.append(headers)
for cell in ws[1]:
    cell.fill = HEADER_FILL
    cell.font = Font(bold=True)

col_widths = {"A":20,"B":28,"C":8,"D":16,"E":10,"F":20,"G":35,"H":28,"I":25,"J":55,"K":38,"L":50,"M":50}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

for g in game_rows:
    row = [g["genre"],g["game"],g["year"],g["revenue"],g["review"],g["team_size"],g["trailer"],g["ui"],g["art_style"],g["feature"],g["scope"],g["content"],g["replayability"]]
    ws.append(row)
    xl_row = ws.max_row
    fill = TIER_FILLS.get(g["tier"])
    if fill:
        for col in range(1, 14):
            ws.cell(xl_row, col).fill = fill
    for col in range(7, 14):
        ws.cell(xl_row, col).alignment = Alignment(wrap_text=True, vertical="top")

ws.freeze_panes = "A2"

output_path = r"C:\Organized Files\My Game Asset\Game-Research\gameplay-review-roguelike.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")

snapshot_dir = Path(r"C:\Organized Files\My Game Asset\Game-Research\snapshots\roguelike")
snapshot_dir.mkdir(parents=True, exist_ok=True)
snapshot_path = snapshot_dir / f"{date.today().isoformat()}.xlsx"
shutil.copy2(output_path, snapshot_path)
print(f"Snapshot: {snapshot_path}")

log_path = Path(r"C:\Organized Files\My Game Asset\Game-Research\research-log.md")
today_str = date.today().isoformat()
high_count = sum(1 for g in game_rows if g["tier"] == "HIGH")
mid_count  = sum(1 for g in game_rows if g["tier"] == "MID")
fail_count = sum(1 for g in game_rows if g["tier"] == "FAILURE")
entry = f"| {today_str} | Roguelike | {high_count} | {mid_count} | {fail_count} | snapshots/roguelike/{today_str}.xlsx |\n"
if log_path.exists():
    existing = log_path.read_text(encoding="utf-8")
    log_path.write_text(existing + entry, encoding="utf-8")
else:
    log_path.write_text(
        "# Research Log\n\n| Date | Genre | HIGH | MID | FAILURE | Snapshot |\n|------|-------|------|-----|---------|----------|\n" + entry,
        encoding="utf-8",
    )
print(f"research-log.md updated. HIGH={high_count} MID={mid_count} FAILURE={fail_count}")
