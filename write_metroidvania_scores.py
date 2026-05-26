import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import gspread
from google.oauth2.service_account import Credentials
from datetime import date
import time
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

CREDENTIALS_FILE = r"C:\Organized Files\My Game Asset\Game-Research\genre-viability-data-417b9f28c38e.json"
SHEET_URL = "https://docs.google.com/spreadsheets/d/1p7vPIIR8imPAZUZekbIuDW7Qfg0Jfc9xa6eEeaSv5Us/edit?usp=sharing"
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

creds  = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=SCOPES)
client = gspread.Client(auth=creds)
sh     = client.open_by_url(SHEET_URL)

COLOR_HEADER  = {"red": 0.812, "green": 0.886, "blue": 0.953}
COLOR_HIGH    = {"red": 0.851, "green": 0.918, "blue": 0.827}
COLOR_MID     = {"red": 1.0,   "green": 0.949, "blue": 0.800}
COLOR_FAILURE = {"red": 1.0,   "green": 0.902, "blue": 0.902}
COLOR_WHITE   = {"red": 1.0,   "green": 1.0,   "blue": 1.0}

SCORE_RED   = PatternFill("solid", fgColor="FFE6E6")
SCORE_GREEN = PatternFill("solid", fgColor="D9EAD3")
SCORE_MID   = PatternFill("solid", fgColor="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="CFE2F3")
HIGH_FILL   = PatternFill("solid", fgColor="D9EAD3")
MID_FILL    = PatternFill("solid", fgColor="FFF2CC")
FAIL_FILL   = PatternFill("solid", fgColor="FFE6E6")

today = date.today().isoformat()
genre = "Metroidvania"

NA = "N/A"

# Each entry: game, category, store_url, art, store, screenshot_url, ui, trailer_url, trailer, hook,
#             feel, feel_src, content, content_src, replay, replay_src
game_rows = [
    # ===== HIGH =====
    {
        "game": "Hollow Knight",
        "category": "HIGH",
        "store_url": "https://store.steampowered.com/app/367520/Hollow_Knight/",
        "art": 10, "store": 9,
        "screenshot_url": "https://store.steampowered.com/app/367520/Hollow_Knight/",
        "ui": 8,
        "trailer_url": "https://www.youtube.com/watch?v=UAO2urG23S4",
        "trailer": 9, "hook": 9,
        "feel": 10, "feel_src": '"combat feels fast, responsive, and satisfying in a way very few games truly master; the sheer act of playing is an absolute joy" — Forever Classic Games review',
        "content": 10, "content_src": '"112% completion takes about 55 hours; Godhome DLC adds a massive endgame gauntlet" — Steam community discussion',
        "replay": 9, "replay_src": '"Steel Soul permadeath mode + Godhome update skyrocketed replayability; speedrunning community still active 8 years later" — Steam community discussion',
    },
    {
        "game": "Dead Cells",
        "category": "HIGH",
        "store_url": "https://store.steampowered.com/app/588650/Dead_Cells/",
        "art": 8, "store": 9,
        "screenshot_url": "https://store.steampowered.com/app/588650/Dead_Cells/",
        "ui": 8,
        "trailer_url": "https://www.youtube.com/watch?v=HRMqBbcWVV4",
        "trailer": 9, "hook": 8,
        "feel": 10, "feel_src": '"best 2D combat — feels so fluid, responsive, and satisfying; controls make it stand out from the rest" — Metacritic user review',
        "content": 9, "content_src": '"hundreds of weapon synergies; years of DLC expanded the game massively after 1.0; still receiving content in 2025" — Steam review consensus',
        "replay": 10, "replay_src": '"roguelike structure means no two runs feel identical; hundreds of build combinations ensure every death teaches something" — Steam review',
    },
    {
        "game": "Animal Well",
        "category": "HIGH",
        "store_url": "https://store.steampowered.com/app/813230/ANIMAL_WELL/",
        "art": 9, "store": 8,
        "screenshot_url": "https://store.steampowered.com/app/813230/ANIMAL_WELL/",
        "ui": 9,
        "trailer_url": "https://www.youtube.com/watch?v=MYNdFGMVFic",
        "trailer": 8, "hook": 9,
        "feel": 8, "feel_src": '"movement feels precise and deliberate; every interaction feels weighted and meaningful in the absence of combat" — Digital Trends review',
        "content": 7, "content_src": '"5-10 hours for main path; 64 secret eggs extend playtime significantly; community still discovering hidden puzzles" — Steam community discussion',
        "replay": 9, "replay_src": '"rooms packed to brim with secrets; knowledge loop creates strong exploration hook; layers of secrets being found by community years later" — Digital Trends review',
    },
    {
        "game": "Ender Magnolia: Bloom in the Mist",
        "category": "HIGH",
        "store_url": "https://store.steampowered.com/app/2725260/ENDER_MAGNOLIA_Bloom_in_the_Mist/",
        "art": 9, "store": 9,
        "screenshot_url": "https://store.steampowered.com/app/2725260/ENDER_MAGNOLIA_Bloom_in_the_Mist/",
        "ui": 8,
        "trailer_url": "https://www.youtube.com/watch?v=YEi5epBvTpk",
        "trailer": 8, "hook": 7,
        "feel": 8, "feel_src": '"combat is smooth and well set out; tight, fluid and challenging; rewards skill and experimentation with unlimited ability uses" — Chalgyr\'s Game Room review',
        "content": 8, "content_src": '"17-18 hours to 100% complete; full release provides approximately 35 hours of content with all hidden areas" — review site aggregate',
        "replay": 7, "replay_src": '"already feel like playing again immediately after finishing; countless locked paths and secret rooms reward re-exploration" — Steam Deck HQ review',
    },
    {
        "game": "Nine Sols",
        "category": "HIGH",
        "store_url": "https://store.steampowered.com/app/1809540/Nine_Sols/",
        "art": 9, "store": 9,
        "screenshot_url": "https://store.steampowered.com/app/1809540/Nine_Sols/",
        "ui": 8,
        "trailer_url": "https://www.youtube.com/watch?v=lkPahRSNVB4",
        "trailer": 9, "hook": 9,
        "feel": 10, "feel_src": '"perfectly timed parry produces a satisfying clang; parrying and talisman attacks are so satisfying; makes fights feel very back and forth like Sekiro" — Lords of Gaming review',
        "content": 9, "content_src": '"expansive world split into multiple zones rich with secrets; previously inaccessible areas become reachable as Yi gains new traversal abilities" — Lords of Gaming review',
        "replay": 7, "replay_src": '"deep lore and alternate boss approaches reward replay; linear structure limits run variety compared to roguelikes but exploration depth compensates" — PC Gamer review',
    },
    {
        "game": "Blasphemous 2",
        "category": "HIGH",
        "store_url": "https://store.steampowered.com/app/2114740/Blasphemous_2/",
        "art": 9, "store": 8,
        "screenshot_url": "https://store.steampowered.com/app/2114740/Blasphemous_2/",
        "ui": 7,
        "trailer_url": "https://www.youtube.com/watch?v=W-5Ai1L6GaM",
        "trailer": 8, "hook": 8,
        "feel": 7, "feel_src": '"three weapons feel completely different — rapier quick and short-range, flail hits hard in huge arc; action and movement feel solid if a bit safe" — Fextralife review',
        "content": 7, "content_src": '"doesn\'t pack as much content as Hollow Knight but deep combat systems and collectibles give solid value for price" — Forever Classic Games review',
        "replay": 7, "replay_src": '"power-ups and collectibles give reason to revisit previous levels; second ending unlockable; sheer number of secrets rewards thorough players" — review',
    },
    {
        "game": "Pseudoregalia",
        "category": "HIGH",
        "store_url": "https://store.steampowered.com/app/2365810/Pseudoregalia/",
        "art": 6, "store": 7,
        "screenshot_url": "https://store.steampowered.com/app/2365810/Pseudoregalia/",
        "ui": 7,
        "trailer_url": "https://www.youtube.com/watch?v=4bKoS0-cQsc",
        "trailer": 7, "hook": 8,
        "feel": 10, "feel_src": '"movement is smooth as butter; don\'t feel like I can go back to a game with limited movement after this; the game feel was so good" — Metacritic user review',
        "content": 4, "content_src": '"game is about 2-4 hours long; condensed but full metroidvania experience; players want more immediately after finishing" — Steam community',
        "replay": 7, "replay_src": '"highly replayable with different routes; players replay multiple times just to feel the controls and hear the music" — Pixeldie review',
    },
    # ===== MID =====
    {
        "game": "Gestalt: Steam & Cinder",
        "category": "MID",
        "store_url": "https://store.steampowered.com/app/1231990/Gestalt_Steam__Cinder/",
        "art": 8, "store": 8,
        "screenshot_url": "https://store.steampowered.com/app/1231990/Gestalt_Steam__Cinder/",
        "ui": 7,
        "trailer_url": "https://www.youtube.com/watch?v=5ZjzFv04f3E",
        "trailer": 7, "hook": 6,
        "feel": NA, "feel_src": "N/A — insufficient player reviews found in search",
        "content": NA, "content_src": "N/A — insufficient player reviews found in search",
        "replay": NA, "replay_src": "N/A — insufficient player reviews found in search",
    },
    {
        "game": "Moonscars",
        "category": "MID",
        "store_url": "https://store.steampowered.com/app/1374970/Moonscars/",
        "art": 8, "store": 7,
        "screenshot_url": "https://store.steampowered.com/app/1374970/Moonscars/",
        "ui": 6,
        "trailer_url": "https://www.youtube.com/watch?v=FV_MkQ4glLU",
        "trailer": 7, "hook": 6,
        "feel": 6, "feel_src": '"snappy fast paced combat; parry window is tiny; enemies have a single move so parry becomes trivial quickly outside of bosses" — Steam review',
        "content": 5, "content_src": '"beat the game in one sitting; short enough that lack of enemy variety wasn\'t a major issue" — Steam community',
        "replay": 6, "replay_src": '"subsequent playthroughs mandatory to experience all content but don\'t really clear up story confusion" — Metacritic user review',
    },
    {
        "game": "Bo: Path of the Teal Lotus",
        "category": "MID",
        "store_url": "https://store.steampowered.com/app/1614440/B_Path_of_the_Teal_Lotus/",
        "art": 9, "store": 8,
        "screenshot_url": "https://store.steampowered.com/app/1614440/B_Path_of_the_Teal_Lotus/",
        "ui": 7,
        "trailer_url": "https://www.youtube.com/watch?v=LKQHxU9wV_s",
        "trailer": 8, "hook": 7,
        "feel": NA, "feel_src": "N/A — insufficient player reviews found in search",
        "content": NA, "content_src": "N/A — insufficient player reviews found in search",
        "replay": NA, "replay_src": "N/A — insufficient player reviews found in search",
    },
    {
        "game": "Souldiers",
        "category": "MID",
        "store_url": "https://store.steampowered.com/app/1419160/Souldiers/",
        "art": 6, "store": 6,
        "screenshot_url": "https://store.steampowered.com/app/1419160/Souldiers/",
        "ui": 5,
        "trailer_url": "https://www.youtube.com/watch?v=OwpzOGkr-3Y",
        "trailer": 6, "hook": 5,
        "feel": 5, "feel_src": '"controls often feel loose; character crouches unintentionally; sluggish controls cited by roughly 37% of all Steam reviews as a key pain point" — Steam negative review aggregate',
        "content": 8, "content_src": '"one playthrough took 23 hours at 90% completion; 25-65 hours total depending on class and difficulty" — Steam community discussion',
        "replay": 7, "replay_src": '"three different classes with divergent movesets and skill trees; three difficulty settings provide significant replay variation" — ConFreaksAndGeeks review',
    },
    {
        "game": "HAAK",
        "category": "MID",
        "store_url": "https://store.steampowered.com/app/1352930/HAAK/",
        "art": 7, "store": 7,
        "screenshot_url": "https://store.steampowered.com/app/1352930/HAAK/",
        "ui": 6,
        "trailer_url": "https://www.youtube.com/watch?v=Jrk9TKLQ_c0",
        "trailer": 7, "hook": 7,
        "feel": NA, "feel_src": "N/A — insufficient player reviews found in search",
        "content": NA, "content_src": "N/A — insufficient player reviews found in search",
        "replay": NA, "replay_src": "N/A — insufficient player reviews found in search",
    },
    {
        "game": "Elderand",
        "category": "MID",
        "store_url": "https://store.steampowered.com/app/1413660/Elderand/",
        "art": 7, "store": 6,
        "screenshot_url": "https://store.steampowered.com/app/1413660/Elderand/",
        "ui": 6,
        "trailer_url": "https://www.youtube.com/watch?v=ZuLm3-kMEA8",
        "trailer": 6, "hook": 5,
        "feel": NA, "feel_src": "N/A — insufficient player reviews found in search",
        "content": NA, "content_src": "N/A — insufficient player reviews found in search",
        "replay": NA, "replay_src": "N/A — insufficient player reviews found in search",
    },
    {
        "game": "Lone Fungus",
        "category": "MID",
        "store_url": "https://store.steampowered.com/app/1674780/Lone_Fungus/",
        "art": 7, "store": 7,
        "screenshot_url": "https://store.steampowered.com/app/1674780/Lone_Fungus/",
        "ui": 6,
        "trailer_url": "https://www.youtube.com/watch?v=fJHOoWRH0Ds",
        "trailer": 6, "hook": 6,
        "feel": NA, "feel_src": "N/A — insufficient player reviews found in search",
        "content": NA, "content_src": "N/A — insufficient player reviews found in search",
        "replay": NA, "replay_src": "N/A — insufficient player reviews found in search",
    },
    # ===== FAILURE =====
    {
        "game": "The Last Case of Benedict Fox",
        "category": "FAILURE",
        "store_url": "https://store.steampowered.com/app/2023360/The_Last_Case_of_Benedict_Fox_Definitive_Edition/",
        "art": 8, "store": 7,
        "screenshot_url": "https://store.steampowered.com/app/2023360/The_Last_Case_of_Benedict_Fox_Definitive_Edition/",
        "ui": 5,
        "trailer_url": "https://www.youtube.com/watch?v=kRNBnkk-Bos",
        "trailer": 7, "hook": 6,
        "feel": 4, "feel_src": '"rough around the edges on controls; map and pause menu delayed and not smooth; coming off Hollow Knight this is gonna feel bad" — Steam review',
        "content": 6, "content_src": '"substantial content but missing map markers and poor quality-of-life features make navigation a chore rather than a pleasure" — Steam review',
        "replay": 4, "replay_src": '"Game Pass launched it without building a community; 31 concurrent players months after launch; nobody returns after finishing" — Steambase player data',
    },
    {
        "game": "Curse of the Sea Rats",
        "category": "FAILURE",
        "store_url": "https://store.steampowered.com/app/1453900/Curse_of_the_Sea_Rats/",
        "art": 8, "store": 7,
        "screenshot_url": "https://store.steampowered.com/app/1453900/Curse_of_the_Sea_Rats/",
        "ui": 5,
        "trailer_url": "https://www.youtube.com/watch?v=M_KtSwPGhZ4",
        "trailer": 7, "hook": 6,
        "feel": NA, "feel_src": "N/A — insufficient player reviews found in search",
        "content": NA, "content_src": "N/A — insufficient player reviews found in search",
        "replay": NA, "replay_src": "N/A — insufficient player reviews found in search",
    },
    {
        "game": "Tales of Kenzera: ZAU",
        "category": "FAILURE",
        "store_url": "https://store.steampowered.com/app/2316580/Tales_of_Kenzera_ZAU/",
        "art": 9, "store": 8,
        "screenshot_url": "https://store.steampowered.com/app/2316580/Tales_of_Kenzera_ZAU/",
        "ui": 7,
        "trailer_url": "https://www.youtube.com/watch?v=fE3hAC60f_s",
        "trailer": 8, "hook": 8,
        "feel": 6, "feel_src": '"combat remains more of a burden than a pleasure for much of the game; only starts to take off late when complexity increases — far too late" — OpenCritic review aggregate',
        "content": 6, "content_src": '"8-10 hour experience with very little replay value; short and sweet but feels thin for the price" — Steam review',
        "replay": 4, "replay_src": '"very little replay value once finished; shaman abilities add minimal replayability; playerbase evaporated after PS Plus launch" — Steam review',
    },
    {
        "game": "INAYAH - Life after Gods",
        "category": "FAILURE",
        "store_url": "https://store.steampowered.com/app/3199390/INAYAH__Life_after_Gods/",
        "art": 7, "store": 6,
        "screenshot_url": "https://store.steampowered.com/app/3199390/INAYAH__Life_after_Gods/",
        "ui": 6,
        "trailer_url": "https://www.youtube.com/watch?v=GaHoJ3JGnVE",
        "trailer": 6, "hook": 6,
        "feel": NA, "feel_src": "N/A — too few reviews (~205) to find reliable citations in search",
        "content": NA, "content_src": "N/A — too few reviews (~205) to find reliable citations in search",
        "replay": NA, "replay_src": "N/A — too few reviews (~205) to find reliable citations in search",
    },
    {
        "game": "Vigil: The Longest Night",
        "category": "FAILURE",
        "store_url": "https://store.steampowered.com/app/1348260/Vigil_The_Longest_Night/",
        "art": 8, "store": 7,
        "screenshot_url": "https://store.steampowered.com/app/1348260/Vigil_The_Longest_Night/",
        "ui": 5,
        "trailer_url": "https://www.youtube.com/watch?v=yFbYLVT63Js",
        "trailer": 7, "hook": 7,
        "feel": NA, "feel_src": "N/A — insufficient player reviews found in search",
        "content": NA, "content_src": "N/A — insufficient player reviews found in search",
        "replay": NA, "replay_src": "N/A — insufficient player reviews found in search",
    },
    {
        "game": "Aeterna Noctis",
        "category": "FAILURE",
        "store_url": "https://store.steampowered.com/app/1517970/Aeterna_Noctis/",
        "art": 8, "store": 7,
        "screenshot_url": "https://store.steampowered.com/app/1517970/Aeterna_Noctis/",
        "ui": 6,
        "trailer_url": "https://www.youtube.com/watch?v=M3i_U_hSLsM",
        "trailer": 7, "hook": 6,
        "feel": NA, "feel_src": "N/A — insufficient player reviews found in search",
        "content": NA, "content_src": "N/A — insufficient player reviews found in search",
        "replay": NA, "replay_src": "N/A — insufficient player reviews found in search",
    },
    {
        "game": "Pronty",
        "category": "FAILURE",
        "store_url": "https://store.steampowered.com/app/1481560/Pronty/",
        "art": 7, "store": 5,
        "screenshot_url": "https://store.steampowered.com/app/1481560/Pronty/",
        "ui": 6,
        "trailer_url": "https://www.youtube.com/watch?v=Jvz17OC2iD4",
        "trailer": 5, "hook": 6,
        "feel": NA, "feel_src": "N/A — too few reviews (~100) to find reliable citations in search",
        "content": NA, "content_src": "N/A — too few reviews (~100) to find reliable citations in search",
        "replay": NA, "replay_src": "N/A — too few reviews (~100) to find reliable citations in search",
    },
]

def numeric(v):
    return isinstance(v, (int, float))

def calc_total(g):
    keys = ["art","store","ui","trailer","hook","feel","content","replay"]
    return sum(g[k] for k in keys if numeric(g.get(k)))

def calc_flag(g):
    if g["category"] == "HIGH":
        return "STRONG"
    if g["category"] == "FAILURE":
        return "LOW"
    total = calc_total(g)
    return "DECENT" if total >= 50 else "WEAK"

# =================== EXCEL ===================
wb = openpyxl.Workbook()
ws_xl = wb.active
ws_xl.title = "Quality Scores"

headers = [
    "Game", "Category",
    "Store URL", "Art", "Store Score",
    "Screenshot URL", "UI",
    "Trailer URL", "Trailer", "Hook",
    "Game Feel", "Feel Source",
    "Content", "Content Source",
    "Replayability", "Replay Source",
    "Total", "Flag"
]
ws_xl.append(headers)
for cell in ws_xl[1]:
    cell.fill = HEADER_FILL
    cell.font = Font(bold=True)

col_widths = {
    "A": 30, "B": 10, "C": 15, "D": 7, "E": 10,
    "F": 15, "G": 6, "H": 15, "I": 7, "J": 7,
    "K": 10, "L": 50, "M": 8, "N": 50, "O": 12, "P": 50,
    "Q": 7, "R": 10
}
for col_letter, width in col_widths.items():
    ws_xl.column_dimensions[col_letter].width = width

def score_fill_xl(val):
    if not numeric(val):
        return None
    if val <= 4:
        return SCORE_RED
    if val >= 7:
        return SCORE_GREEN
    return SCORE_MID

for g in game_rows:
    total = calc_total(g)
    flag  = calc_flag(g)
    row = [
        g["game"], g["category"],
        g["store_url"], g["art"], g["store"],
        g["screenshot_url"], g["ui"],
        g["trailer_url"], g["trailer"], g["hook"],
        g["feel"], g["feel_src"],
        g["content"], g["content_src"],
        g["replay"], g["replay_src"],
        total, flag,
    ]
    ws_xl.append(row)
    xl_row = ws_xl.max_row

    tier_fill = {"HIGH": HIGH_FILL, "MID": MID_FILL, "FAILURE": FAIL_FILL}.get(g["category"])
    if tier_fill:
        ws_xl.cell(xl_row, 1).fill = tier_fill
        ws_xl.cell(xl_row, 2).fill = tier_fill

    for col_idx, url_key in [(3, "store_url"), (6, "screenshot_url"), (8, "trailer_url")]:
        url = g.get(url_key, "")
        if url:
            cell = ws_xl.cell(xl_row, col_idx)
            cell.hyperlink = url
            cell.value = "Link"
            cell.font = Font(color="0563C1", underline="single")

    score_cols = {4: g["art"], 5: g["store"], 7: g["ui"], 9: g["trailer"],
                  10: g["hook"], 11: g["feel"], 13: g["content"], 15: g["replay"]}
    for col_idx, score_val in score_cols.items():
        fill = score_fill_xl(score_val)
        if fill:
            ws_xl.cell(xl_row, col_idx).fill = fill

output_path = r"C:\Organized Files\My Game Asset\Game-Research\gameplay-review-metroidvania.xlsx"
wb.save(output_path)
print(f"Saved Excel: {output_path}")

# =================== GOOGLE SHEET ===================
HEADERS_QUALITY = [
    "Genre", "Game", "Category", "Date",
    "Art", "Store Page", "Trailer", "UI", "Hook",
    "Game Feel", "Feel Source",
    "Content", "Content Source",
    "Replayability", "Replay Source",
    "Total", "Flag"
]

try:
    ws_gs = sh.worksheet("Quality Scores")
except:
    ws_gs = sh.add_worksheet(title="Quality Scores", rows=1000, cols=20)
    ws_gs.append_row(HEADERS_QUALITY)
    ws_gs.format("A1:Q1", {"backgroundColor": COLOR_HEADER, "textFormat": {"bold": True}})

sheet_rows = []
for g in game_rows:
    total = calc_total(g)
    flag  = calc_flag(g)
    sheet_rows.append([
        genre, g["game"], g["category"], today,
        g["art"], g["store"], g["trailer"], g["ui"], g["hook"],
        g["feel"], g["feel_src"],
        g["content"], g["content_src"],
        g["replay"], g["replay_src"],
        total, flag,
    ])

ws_gs.append_rows(sheet_rows, value_input_option="USER_ENTERED")
print(f"Appended {len(sheet_rows)} rows to Quality Scores tab.")

all_vals = ws_gs.get_all_values()
total_rows = len(all_vals)
start_row = total_rows - len(sheet_rows) + 1

format_requests = []
for i, row_data in enumerate(sheet_rows, start=start_row):
    category = row_data[2]
    color = {"HIGH": COLOR_HIGH, "MID": COLOR_MID, "FAILURE": COLOR_FAILURE}.get(category, COLOR_WHITE)
    format_requests.append({
        "repeatCell": {
            "range": {
                "sheetId": ws_gs.id,
                "startRowIndex": i - 1,
                "endRowIndex": i,
                "startColumnIndex": 0,
                "endColumnIndex": 17,
            },
            "cell": {"userEnteredFormat": {"backgroundColor": color}},
            "fields": "userEnteredFormat.backgroundColor",
        }
    })
    if len(format_requests) >= 20:
        sh.batch_update({"requests": format_requests})
        format_requests = []
        time.sleep(1.2)

if format_requests:
    sh.batch_update({"requests": format_requests})

print("Colors applied.")
print(f"Done. Written {len(sheet_rows)} quality score rows for '{genre}'.")
